"""
app.py — ระบบจัดการงานลูกค้า v10 (Fixed)
Bugs fixed:
  1. /model route: was creating empty task on every GET → now just renders template
  2. get_qr_image(): returned 'promptpay.ext' → now returns 'qr/promptpay.ext'
  3. upload_slip: read 'slip_file' field → now reads 'slip'
  4. payment(): passed slip= single object → now passes slips= full list
  5. ticket.html: used wrong field names → fixed to use correct ticket fields
  6. todos.json not initialized → added to _init loop
  7. .env encoding: auto-fix UTF-8 on startup
"""
import json, os, uuid, secrets as _secrets, io as _io, zipfile, tempfile, shutil, smtplib, time, hmac, hashlib, threading, sqlite3
from contextlib import contextmanager
from datetime import datetime, date, timedelta
from functools import wraps
from collections import defaultdict
from email.message import EmailMessage

from flask import (Flask, render_template, request, redirect, url_for,
                   session, jsonify, make_response, send_from_directory, abort)
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash

from promptpay import generate_promptpay_payload
from queue_manager import (
    read_queue, read_calendar, write_calendar,
    sync_queue, reorder_queue, set_task_estimate,
    get_queue_with_tasks, yearly_analytics,
    add_custom_date, remove_custom_date, update_calendar_settings,
    working_days_count, MONTH_TH
)

try:
    from pdf_generator import generate_order_pdf, generate_spec_sheet
    PDF_ENABLED = True
except Exception:
    PDF_ENABLED = False
    generate_spec_sheet = None

try:
    from line_handler import handle_events, verify_signature
    LINE_ENABLED = True
except Exception:
    LINE_ENABLED = False

# ── FIX 7: Auto-fix .env encoding on startup (Windows cp1252 → UTF-8) ──────────
def _fix_env_encoding():
    env_path = os.path.join(os.path.dirname(__file__), '.env')
    if not os.path.exists(env_path): return
    for enc in ('utf-8', 'utf-8-sig', 'cp1252', 'tis-620', 'latin-1'):
        try:
            with open(env_path, 'r', encoding=enc) as f: text = f.read()
            with open(env_path, 'w', encoding='utf-8') as f: f.write(text)
            break
        except (UnicodeDecodeError, LookupError): continue
_fix_env_encoding()

# ── App setup ──────────────────────────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'change-this-secret-in-production')

from werkzeug.middleware.proxy_fix import ProxyFix
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)

PREFERRED_SCHEME = os.environ.get('PREFERRED_SCHEME', '')
PROJECT_DIR      = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER    = os.path.join(PROJECT_DIR, 'uploads')
QR_FOLDER        = os.path.join(UPLOAD_FOLDER, 'qr')
SLIP_FOLDER      = os.path.join(UPLOAD_FOLDER, 'slips')
PRODUCT_IMG_FOLDER = os.path.join(UPLOAD_FOLDER, 'products')
ALLOWED_IMG      = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
MODEL_3D_FOLDER  = os.path.join(UPLOAD_FOLDER, '3d_models')
GALLERY_FOLDER   = os.path.join(UPLOAD_FOLDER, 'gallery')
CUSTOM_ORDER_FOLDER = os.path.join(UPLOAD_FOLDER, 'custom_orders')
BACKUP_FOLDER    = os.path.join(PROJECT_DIR, 'backups')
ALLOWED_3D       = {'stl', 'obj', 'step', 'stp', '3mf', 'iges', 'igs', 'f3d', 'blend', 'fbx', 'zip'}
ALLOWED_ORDER_FILES = ALLOWED_IMG | {'pdf', 'ai', 'psd', 'svg', 'zip', 'doc', 'docx', 'xls', 'xlsx'}

for d in [UPLOAD_FOLDER, QR_FOLDER, SLIP_FOLDER, PRODUCT_IMG_FOLDER, MODEL_3D_FOLDER, GALLERY_FOLDER, CUSTOM_ORDER_FOLDER, BACKUP_FOLDER]:
    os.makedirs(d, exist_ok=True)

PROMPTPAY_PHONE  = os.environ.get('PROMPTPAY_PHONE', '0812345678')
COMPANY_NAME     = os.environ.get('COMPANY_NAME', 'ระบบจัดการงานลูกค้า')
STAMPS_TO_REWARD = 10
JSON_LOCK_TIMEOUT = 8
_JSON_THREAD_LOCK = threading.RLock()
_LOGIN_ATTEMPTS = {}

try:
    import bcrypt as _bcrypt
except Exception:
    _bcrypt = None

def hash_password(password):
    if _bcrypt:
        return 'bcrypt$' + _bcrypt.hashpw(password.encode('utf-8'), _bcrypt.gensalt()).decode('utf-8')
    return generate_password_hash(password)

def verify_password(stored, password):
    stored = stored or ''
    if stored.startswith('bcrypt$') and _bcrypt:
        return _bcrypt.checkpw(password.encode('utf-8'), stored[7:].encode('utf-8'))
    if stored.startswith(('scrypt:', 'pbkdf2:', 'argon2:')):
        return check_password_hash(stored, password)
    legacy_sha = hashlib.sha256(password.encode()).hexdigest()
    return hmac.compare_digest(stored, password) or hmac.compare_digest(stored, legacy_sha)

def password_needs_upgrade(stored):
    return not (stored or '').startswith(('bcrypt$', 'scrypt:', 'pbkdf2:', 'argon2:'))

@contextmanager
def json_file_lock(path):
    lock_path = f'{path}.lock'
    deadline = time.time() + JSON_LOCK_TIMEOUT
    with _JSON_THREAD_LOCK:
        fd = None
        while fd is None:
            try:
                fd = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_RDWR)
                os.write(fd, str(os.getpid()).encode('ascii'))
            except FileExistsError:
                if time.time() > deadline:
                    raise TimeoutError(f'Could not lock {path}')
                time.sleep(0.05)
        try:
            yield
        finally:
            try:
                os.close(fd)
            finally:
                try: os.unlink(lock_path)
                except FileNotFoundError: pass

def csrf_token():
    token = session.get('_csrf_token')
    if not token:
        token = _secrets.token_urlsafe(32)
        session['_csrf_token'] = token
    return token

@app.context_processor
def inject_globals():
    try: cc = cart_count()
    except: cc = 0
    return dict(cart_count=cc, company_name=COMPANY_NAME, csrf_token=csrf_token)

@app.before_request
def protect_post_requests():
    maybe_auto_backup()
    if request.method not in {'POST', 'PUT', 'PATCH', 'DELETE'}:
        return
    if request.path == '/webhook':
        return
    sent = request.form.get('csrf_token') or request.headers.get('X-CSRFToken') or request.headers.get('X-CSRF-Token')
    if request.is_json and not sent:
        data = request.get_json(silent=True) or {}
        sent = data.get('csrf_token')
    if not sent or not hmac.compare_digest(sent, csrf_token()):
        return jsonify({'status':'error','error':'csrf'}), 400

# ── Data init ──────────────────────────────────────────────────────────────────
def _init(path, default):
    if not os.path.exists(path):
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(default, f, ensure_ascii=False)

# FIX 6: Added todos.json to init list
for p, d in [
    ('tasks.json', []),
    ('users.json', {'admin': 'admin123'}),
    ('stamps.json', {}),
    ('tickets.json', {}),
    ('slips.json', {}),
    ('products.json', []),
    ('orders_cart.json', {}),
    ('sn_counter.json', {'last_sn': 0}),
    ('todos.json', []),          # ← FIX 6
    ('events.json', []),
    ('notifications.json', []),
    ('gallery.json', []),
    ('reviews.json', []),
    ('coupons.json', []),
    ('invoices.json', {'last_no': 0, 'items': {}}),
]:
    _init(p, d)

def _r(p, default):
    try:
        with json_file_lock(p):
            with open(p, 'r', encoding='utf-8') as f: return json.load(f)
    except: return default

def _w(p, data):
    with json_file_lock(p):
        tmp = f'{p}.{os.getpid()}.tmp'
        with open(tmp, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        os.replace(tmp, p)

read_tasks    = lambda: _r('tasks.json', [])
write_tasks   = lambda d: _w('tasks.json', d)
read_users    = lambda: _r('users.json', {'admin': 'admin123'})
write_users   = lambda d: _w('users.json', d)
read_stamps   = lambda: _r('stamps.json', {})
write_stamps  = lambda d: _w('stamps.json', d)
read_tickets  = lambda: _r('tickets.json', {})
write_tickets = lambda d: _w('tickets.json', d)
read_slips    = lambda: _r('slips.json', {})
write_slips   = lambda d: _w('slips.json', d)
read_products = lambda: _r('products.json', [])
write_products= lambda d: _w('products.json', d)
read_todos    = lambda: _r('todos.json', [])
write_todos   = lambda d: _w('todos.json', d)
read_events   = lambda: _r('events.json', [])
write_events  = lambda d: _w('events.json', d)
read_notifications  = lambda: _r('notifications.json', [])
write_notifications = lambda d: _w('notifications.json', d)
read_gallery  = lambda: _r('gallery.json', [])
write_gallery = lambda d: _w('gallery.json', d)
read_reviews  = lambda: _r('reviews.json', [])
write_reviews = lambda d: _w('reviews.json', d)
read_coupons  = lambda: _r('coupons.json', [])
write_coupons = lambda d: _w('coupons.json', d)
read_invoices = lambda: _r('invoices.json', {'last_no': 0, 'items': {}})
write_invoices= lambda d: _w('invoices.json', d)
read_sn       = lambda: _r('sn_counter.json', {'last_sn': 0})
write_sn      = lambda d: _w('sn_counter.json', d)

STATUS_FLOW = ['pending', 'quoted', 'approved', 'inprogress', 'printing', 'postprocessing', 'qc', 'ready', 'delivered', 'completed', 'cancelled']
STATUS_LABELS = {
    'pending': 'Pending', 'quoted': 'Quoted', 'approved': 'Approved',
    'inprogress': 'In Progress', 'printing': 'Printing',
    'postprocessing': 'Post-processing', 'qc': 'QC',
    'ready': 'Ready for Pickup', 'delivered': 'Delivered',
    'completed': 'Completed', 'cancelled': 'Cancelled',
}
STATUS_PROGRESS = {
    'pending': 10, 'quoted': 20, 'approved': 30, 'inprogress': 45,
    'printing': 60, 'postprocessing': 72, 'qc': 82, 'ready': 92,
    'delivered': 97, 'completed': 100, 'cancelled': 0,
}

def _now():
    return datetime.now().isoformat()

def find_task(tasks, task_id):
    return next((t for t in tasks if t.get('id') == task_id), None)

def add_event(task_id, action, note='', actor=None, meta=None):
    events = read_events()
    ev = {
        'id': uuid.uuid4().hex[:12],
        'task_id': task_id,
        'action': action,
        'note': note,
        'actor': actor or session.get('username') or session.get('customer_name') or 'system',
        'meta': meta or {},
        'created_at': _now(),
    }
    events.insert(0, ev)
    write_events(events)
    return ev

def events_for_task(task_id):
    return [e for e in read_events() if e.get('task_id') == task_id]

def log_notification(task_id, channel, target, status, subject='', error=''):
    rows = read_notifications()
    rows.insert(0, {
        'id': uuid.uuid4().hex[:12], 'task_id': task_id, 'channel': channel,
        'target': target, 'status': status, 'subject': subject, 'error': error,
        'created_at': _now(),
    })
    write_notifications(rows)

def send_email_notification(task, subject, body):
    target = task.get('customer', {}).get('email', '')
    if not target:
        return
    host = os.environ.get('SMTP_HOST', '')
    if not host:
        log_notification(task['id'], 'email', target, 'skipped', subject, 'SMTP_HOST not configured')
        return
    try:
        msg = EmailMessage()
        msg['Subject'] = subject
        msg['From'] = os.environ.get('SMTP_FROM') or os.environ.get('SMTP_USER') or 'noreply@zerphyrus.local'
        msg['To'] = target
        msg.set_content(body)
        port = int(os.environ.get('SMTP_PORT', '587'))
        with smtplib.SMTP(host, port, timeout=5) as smtp:
            if os.environ.get('SMTP_TLS', '1') != '0':
                smtp.starttls()
            user = os.environ.get('SMTP_USER', '')
            if user:
                smtp.login(user, os.environ.get('SMTP_PASSWORD') or os.environ.get('SMTP_PASS', ''))
            smtp.send_message(msg)
        log_notification(task['id'], 'email', target, 'sent', subject)
    except Exception as e:
        log_notification(task['id'], 'email', target, 'failed', subject, str(e))

def send_line_admin_notification(task, text):
    token = os.environ.get('LINE_CHANNEL_ACCESS_TOKEN', '')
    admin_id = os.environ.get('ADMIN_LINE_USER_ID', '')
    if not token or not admin_id:
        log_notification(task['id'], 'line', admin_id or '-', 'skipped', 'LINE admin notice', 'LINE not configured')
        return
    try:
        import requests
        resp = requests.post(
            'https://api.line.me/v2/bot/message/push',
            headers={'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'},
            json={'to': admin_id, 'messages': [{'type': 'text', 'text': text}]},
            timeout=3,
        )
        log_notification(task['id'], 'line', admin_id, 'sent' if resp.ok else 'failed', 'LINE admin notice', resp.text[:300])
    except Exception as e:
        log_notification(task['id'], 'line', admin_id, 'failed', 'LINE admin notice', str(e))

def send_sms_notification(task, text):
    phone = ''.join(c for c in task.get('customer', {}).get('phone', '') if c.isdigit())
    if not phone:
        return
    sid = os.environ.get('TWILIO_ACCOUNT_SID', '')
    token = os.environ.get('TWILIO_AUTH_TOKEN', '')
    sender = os.environ.get('TWILIO_FROM', '')
    if not (sid and token and sender):
        log_notification(task['id'], 'sms', phone, 'skipped', 'SMS status notice', 'Twilio not configured')
        return
    if phone.startswith('0'):
        phone = '+66' + phone[1:]
    elif not phone.startswith('+'):
        phone = '+' + phone
    try:
        import requests
        resp = requests.post(
            f'https://api.twilio.com/2010-04-01/Accounts/{sid}/Messages.json',
            data={'From': sender, 'To': phone, 'Body': text},
            auth=(sid, token),
            timeout=5,
        )
        log_notification(task['id'], 'sms', phone, 'sent' if resp.ok else 'failed', 'SMS status notice', resp.text[:300])
    except Exception as e:
        log_notification(task['id'], 'sms', phone, 'failed', 'SMS status notice', str(e))

def notify_status_change(task, old_status, new_status):
    title = task.get('sn') or task.get('title', task['id'])
    subject = f'Order {title}: {STATUS_LABELS.get(new_status, new_status)}'
    body = f"Your order status changed from {STATUS_LABELS.get(old_status, old_status)} to {STATUS_LABELS.get(new_status, new_status)}."
    send_email_notification(task, subject, body)
    send_sms_notification(task, body)

def _num(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default

def _int(value, default=0):
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default

def _deadline_days(deadline):
    try:
        if not deadline:
            return None
        return (date.fromisoformat(deadline) - date.today()).days
    except Exception:
        return None

def calculate_3d_price(specs, overrides=None):
    overrides = overrides or {}
    material = specs.get('material') or 'PLA'
    material_cfg = {
        'PLA': {'gram': 3.0, 'density': 1.24}, 'ABS': {'gram': 3.6, 'density': 1.04},
        'PETG': {'gram': 3.8, 'density': 1.27}, 'TPU': {'gram': 5.5, 'density': 1.21},
        'Resin': {'gram': 7.0, 'density': 1.12}, 'Nylon': {'gram': 6.5, 'density': 1.15},
        'ASA': {'gram': 5.0, 'density': 1.07}, 'CF-PLA': {'gram': 6.0, 'density': 1.30},
    }
    quality = {'draft': .85, 'standard': 1.0, 'fine': 1.35, 'ultra': 1.8}
    finish = {'as_printed': 0, 'sanded': 80, 'polished': 140, 'painted': 220}
    support = {'none': {'fee': 0, 'waste': 0}, 'auto': {'fee': 40, 'waste': .12},
               'minimal': {'fee': 60, 'waste': .18}, 'full': {'fee': 120, 'waste': .35}}
    qty = max(1, _int(specs.get('quantity'), 1))
    sx = _num(specs.get('size_x'))
    sy = _num(specs.get('size_y'))
    sz = _num(specs.get('size_z'))
    volume_cm3 = float(specs.get('volume_cm3') or 0) or max(0, sx * sy * sz / 1000)
    infill = max(5, min(100, _num(specs.get('infill'), 20))) / 100
    cfg = material_cfg.get(material, material_cfg['PLA'])
    support_cfg = support.get(specs.get('support'), support['auto'])
    billable_volume = volume_cm3 * (0.28 + infill) * (1 + support_cfg['waste'])
    material_weight_g = billable_volume * cfg['density'] * qty
    material_cost = material_weight_g * cfg['gram']
    machine_hours = max(.4, (volume_cm3 * qty / 22) * quality.get(specs.get('quality'), 1.0))
    machine_cost = machine_hours * _num(overrides.get('machine_hour_rate'), 45)
    fees = finish.get(specs.get('finish'), 0) + support_cfg['fee']
    rush_multiplier = 1.0
    days_left = _deadline_days(specs.get('deadline'))
    if days_left is not None and days_left <= 3:
        rush_multiplier = 1.35
    elif days_left is not None and days_left <= 7:
        rush_multiplier = 1.15
    minimum = float(overrides.get('minimum') or 150)
    discount = float(overrides.get('discount') or 0)
    subtotal = (material_cost + machine_cost + fees) * rush_multiplier
    amount = max(minimum, subtotal - discount)
    deposit_percent = float(overrides.get('deposit_percent') or 50)
    gaps = []
    if not volume_cm3:
        gaps.append('ไม่มี volume จาก STL หรือขนาดชิ้นงาน')
    if not specs.get('material'):
        gaps.append('ยังไม่ได้เลือกวัสดุ')
    if not specs.get('quantity'):
        gaps.append('ยังไม่ได้ระบุจำนวน')
    confidence = 'high' if volume_cm3 and specs.get('material') else ('medium' if sx and sy and sz else 'low')
    return {
        'volume_cm3': round(volume_cm3, 2), 'material_rate': cfg['gram'],
        'material_weight_g': round(material_weight_g, 1),
        'machine_hours': round(machine_hours, 2),
        'material_cost': round(material_cost, 2),
        'machine_cost': round(machine_cost, 2),
        'finish_support_fee': round(fees, 2),
        'rush_multiplier': rush_multiplier,
        'subtotal': round(subtotal, 2), 'discount': round(discount, 2),
        'amount': round(amount, 2), 'deposit_percent': deposit_percent,
        'deposit_amount': round(amount * deposit_percent / 100, 2),
        'balance_amount': round(amount - (amount * deposit_percent / 100), 2),
        'confidence': confidence, 'pricing_gaps': gaps,
    }

def calculate_custom_order_price(specs, overrides=None):
    overrides = overrides or {}
    service = specs.get('service_type') or 'custom'
    qty = max(1, _int(specs.get('quantity'), 1))
    base_rates = {
        'design': 450, 'laser': 180, 'cnc': 650, 'print': 120,
        'assembly': 300, 'repair': 250, 'custom': 300,
    }
    complexity_mult = {'simple': 1.0, 'standard': 1.35, 'complex': 1.9}
    finish_fee = {'none': 0, 'basic': 80, 'premium': 220}
    service_rate = _num(overrides.get('service_rate'), base_rates.get(service, 300))
    complexity = specs.get('complexity') or 'standard'
    width = _num(specs.get('width_mm'))
    height = _num(specs.get('height_mm'))
    depth = _num(specs.get('depth_mm'))
    area_cm2 = max(0, width * height / 100)
    volume_cm3 = max(0, width * height * depth / 1000)
    size_factor = max(1, (area_cm2 / 120) or (volume_cm3 / 80) or 1)
    labor_hours = max(_num(specs.get('labor_hours'), 0), size_factor * complexity_mult.get(complexity, 1.35))
    rush_multiplier = 1.0
    days_left = _deadline_days(specs.get('deadline'))
    if days_left is not None and days_left <= 3:
        rush_multiplier = 1.35
    elif days_left is not None and days_left <= 7:
        rush_multiplier = 1.15
    subtotal = (service_rate * labor_hours + finish_fee.get(specs.get('finish_level'), 80)) * qty * rush_multiplier
    minimum = _num(overrides.get('minimum'), 250)
    discount = _num(overrides.get('discount'), 0)
    amount = max(minimum, subtotal - discount)
    deposit_percent = _num(overrides.get('deposit_percent'), 50)
    gaps = []
    if not specs.get('service_type'):
        gaps.append('ยังไม่ได้เลือกประเภทงาน')
    if not specs.get('width_mm') and not specs.get('height_mm') and not specs.get('depth_mm'):
        gaps.append('ไม่มีขนาดสำหรับประเมินต้นทุน')
    if not specs.get('reference_files'):
        gaps.append('ไม่มีไฟล์อ้างอิง')
    return {
        'service_rate': round(service_rate, 2), 'labor_hours': round(labor_hours, 2),
        'area_cm2': round(area_cm2, 2), 'volume_cm3': round(volume_cm3, 2),
        'rush_multiplier': rush_multiplier, 'subtotal': round(subtotal, 2),
        'discount': round(discount, 2), 'amount': round(amount, 2),
        'deposit_percent': deposit_percent,
        'deposit_amount': round(amount * deposit_percent / 100, 2),
        'balance_amount': round(amount - (amount * deposit_percent / 100), 2),
        'confidence': 'medium' if not gaps else 'low',
        'pricing_gaps': gaps,
    }

def stl_volume_cm3(path):
    try:
        from stl import mesh
        m = mesh.Mesh.from_file(path)
        volume_mm3, _, _ = m.get_mass_properties()
        return round(abs(float(volume_mm3)) / 1000, 2)
    except Exception:
        return 0

def optimize_image(path, max_size=(1600, 1600), thumb=False):
    try:
        from PIL import Image, ImageOps
        with Image.open(path) as im:
            im = ImageOps.exif_transpose(im)
            im.thumbnail((360, 360) if thumb else max_size)
            if im.mode not in ('RGB', 'RGBA'):
                im = im.convert('RGB')
            im.save(path, optimize=True, quality=82)
    except Exception:
        pass

def save_optimized_upload(file, folder, filename):
    path = os.path.join(folder, filename)
    file.save(path)
    optimize_image(path)
    base, ext = os.path.splitext(filename)
    thumb_name = f'{base}_thumb{ext}'
    thumb_path = os.path.join(folder, thumb_name)
    try:
        shutil.copyfile(path, thumb_path)
        optimize_image(thumb_path, thumb=True)
    except Exception:
        thumb_name = ''
    return filename, thumb_name

def create_backup_zip(auto=False):
    stamp = datetime.now().strftime('%Y%m%d_%H%M')
    fname = f"zerphyrus_{'auto_' if auto else ''}{stamp}.zip"
    path = os.path.join(BACKUP_FOLDER, fname)
    with zipfile.ZipFile(path, 'w', zipfile.ZIP_DEFLATED) as z:
        for name in [f for f in os.listdir('.') if f.endswith('.json')]:
            z.write(name, name)
        for root, _, files in os.walk(UPLOAD_FOLDER):
            for fn in files:
                fp = os.path.join(root, fn)
                z.write(fp, os.path.relpath(fp, PROJECT_DIR))
    return path

def maybe_auto_backup():
    today = date.today().isoformat()
    marker = os.path.join(BACKUP_FOLDER, '.last_auto_backup')
    try:
        last = open(marker, encoding='utf-8').read().strip() if os.path.exists(marker) else ''
        if last != today:
            create_backup_zip(auto=True)
            with open(marker, 'w', encoding='utf-8') as f: f.write(today)
    except Exception:
        pass

def migrate_json_to_sqlite(db_path='zerphyrus.sqlite3'):
    conn = sqlite3.connect(db_path)
    try:
        conn.execute('create table if not exists kv_store (name text primary key, data text not null, updated_at text not null)')
        for name in [f for f in os.listdir('.') if f.endswith('.json')]:
            data = open(name, encoding='utf-8').read()
            conn.execute('replace into kv_store(name,data,updated_at) values (?,?,?)', (name, data, _now()))
        conn.commit()
    finally:
        conn.close()
    return db_path

def build_action_items(tasks, slips):
    today = date.today()
    items = []
    for t in tasks:
        status = t.get('status', 'pending')
        quote = t.get('quote', {})
        if t.get('specs_3d') and quote.get('status') in (None, '', 'draft', 'needed'):
            items.append({'type': 'quote', 'task': t, 'label': 'Needs quote', 'tone': 'amber'})
        for s in slips.get(t['id'], []):
            if s.get('status') == 'pending':
                items.append({'type': 'slip', 'task': t, 'label': 'Slip pending', 'tone': 'green'})
        dl = t.get('deadline')
        if dl and status not in ('completed', 'cancelled', 'delivered'):
            try:
                days = (date.fromisoformat(dl[:10]) - today).days
                if days <= 1:
                    items.append({'type': 'deadline', 'task': t, 'label': f'Deadline {days}d', 'tone': 'red' if days < 0 else 'amber'})
            except Exception:
                pass
        try:
            updated = datetime.fromisoformat(t.get('updatedAt') or t.get('createdAt')).date()
            if status == 'pending' and (today - updated).days >= 1:
                items.append({'type': 'stale', 'task': t, 'label': 'Pending >24h', 'tone': 'amber'})
        except Exception:
            pass
    return items[:12]

def revenue_analytics(slips):
    monthly = defaultdict(float)
    total = 0.0
    approved = 0
    for task_slips in slips.values():
        for s in task_slips:
            if s.get('status') == 'approved':
                try:
                    amt = float(s.get('amount') or 0)
                except Exception:
                    amt = 0
                total += amt; approved += 1
                key = (s.get('verified_at') or s.get('uploaded_at') or '')[:7]
                if key: monthly[key] += amt
    months = sorted(monthly.keys())[-12:]
    target = float(os.environ.get('MONTHLY_REVENUE_TARGET', '0') or 0)
    return {
        'total': round(total, 2), 'approved_slips': approved,
        'month_labels': months, 'month_values': [round(monthly[m], 2) for m in months],
        'target': target,
    }

def crm_summary(tasks):
    customers = read_customers()
    rows = {}
    for phone, c in customers.items():
        rows[phone] = {
            'phone': phone, 'name': c.get('name',''), 'email': c.get('email',''),
            'tags': c.get('tags', []), 'note': c.get('note',''),
            'order_count': 0, 'total_spend': 0.0, 'last_order': '',
        }
    for t in tasks:
        c = t.get('customer', {})
        phone = c.get('phone') or 'unknown'
        row = rows.setdefault(phone, {
            'phone': phone, 'name': c.get('name',''), 'email': c.get('email',''),
            'tags': [], 'note': '', 'order_count': 0, 'total_spend': 0.0, 'last_order': '',
        })
        row['name'] = row['name'] or c.get('name','')
        row['email'] = row['email'] or c.get('email','')
        row['order_count'] += 1
        row['total_spend'] += _num(t.get('quote', {}).get('amount') or t.get('order_total'))
        row['last_order'] = max(row['last_order'], t.get('createdAt',''))
    return sorted(rows.values(), key=lambda r: (r['total_spend'], r['order_count'], r['last_order']), reverse=True)

def invoice_for_task(task):
    invoices = read_invoices()
    items = invoices.setdefault('items', {})
    if task['id'] not in items:
        invoices['last_no'] = invoices.get('last_no', 0) + 1
        amount = float(task.get('quote', {}).get('amount') or task.get('order_total') or 0)
        vat_rate = float(os.environ.get('VAT_RATE', '7') or 0)
        subtotal = round(amount / (1 + vat_rate / 100), 2) if vat_rate else amount
        vat = round(amount - subtotal, 2)
        items[task['id']] = {
            'invoice_no': f"INV-{datetime.now().strftime('%Y%m')}-{invoices['last_no']:04d}",
            'task_id': task['id'], 'subtotal': subtotal, 'vat_rate': vat_rate,
            'vat': vat, 'total': round(amount, 2), 'created_at': _now(),
        }
        write_invoices(invoices)
    return items[task['id']]

def next_sn() -> str:
    counter = read_sn(); counter['last_sn'] = counter.get('last_sn', 0) + 1; write_sn(counter)
    return f"ORD-{datetime.now().strftime('%Y%m')}-{counter['last_sn']:04d}"

def backfill_sn():
    tasks = read_tasks(); counter = read_sn(); changed = False
    for t in reversed(tasks):
        if not t.get('sn'):
            counter['last_sn'] = counter.get('last_sn', 0) + 1
            t['sn'] = f"ORD-MIGR-{counter['last_sn']:04d}"; changed = True
    if changed: write_sn(counter); write_tasks(tasks)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_IMG

# ── FIX 2: get_qr_image returns correct subpath 'qr/promptpay.ext' ────────────
def get_qr_image():
    """Return 'qr/promptpay.ext' for URL use — file lives at uploads/qr/."""
    for ext in ALLOWED_IMG:
        path = os.path.join(QR_FOLDER, f'promptpay.{ext}')
        if os.path.exists(path):
            return f'qr/promptpay.{ext}'   # FIX 2
    return None

def add_stamp(phone, name=''):
    s = read_stamps()
    if phone not in s:
        s[phone] = {'stamps': 0, 'total_earned': 0, 'rewards_redeemed': 0, 'name': name}
    s[phone]['stamps'] += 1; s[phone]['total_earned'] += 1; write_stamps(s)

def create_ticket(task):
    code = uuid.uuid4().hex[:8].upper(); t = read_tickets()
    t[code] = {'task_id': task['id'], 'customer_name': task['customer']['name'],
               'customer_phone': task['customer']['phone'], 'task_title': task['title'],
               'status': 'active', 'created_at': datetime.now().isoformat(),
               'checked_in_at': None, 'checked_in_by': None}
    write_tickets(t); return code

def build_analytics(tasks):
    sc = defaultdict(int); pc = defaultdict(int)
    for t in tasks: sc[t['status']] += 1; pc[t.get('priority','medium')] += 1
    today = date.today()
    days = [(today - timedelta(days=i)).isoformat() for i in range(13,-1,-1)]
    dc = defaultdict(int)
    for t in tasks: dc[t['createdAt'][:10]] += 1
    n = len(tasks); c = sc.get('completed', 0)
    return {
        'status_labels': ['รอดำเนินการ','กำลังทำ','เสร็จสิ้น','ยกเลิก'],
        'status_values': [sc.get(k,0) for k in ['pending','inprogress','completed','cancelled']],
        'priority_labels': ['เร่งด่วน','ปานกลาง','ไม่เร่งด่วน'],
        'priority_values': [pc.get(k,0) for k in ['high','medium','low']],
        'day_labels': days, 'day_values': [dc[d] for d in days],
        'completion_rate': round(c/n*100,1) if n else 0,
    }

def slip_status_for_task(task_id):
    slips = read_slips(); task_slips = slips.get(task_id, [])
    return task_slips[-1] if task_slips else None

def slips_for_task(task_id):
    """Return all slips for a task as a list."""
    return read_slips().get(task_id, [])

def pending_slips_count():
    return sum(1 for ts in read_slips().values() for s in ts if s.get('status')=='pending')

def get_webhook_url():
    base = request.url_root.rstrip('/')
    if PREFERRED_SCHEME: base = base.replace('http://', f'{PREFERRED_SCHEME}://')
    elif request.headers.get('X-Forwarded-Proto') == 'https': base = base.replace('http://','https://')
    return f"{base}/webhook"

# ── Admin context ──────────────────────────────────────────────────────────────
def admin_context(tasks_override=None):
    all_tasks = read_tasks(); tasks = tasks_override if tasks_override is not None else all_tasks
    stamps = read_stamps(); tickets = read_tickets(); cal = read_calendar()
    yr = date.today().year; slips = read_slips()
    tasks_with_slip = [{**t, 'slip': slip_status_for_task(t['id'])} for t in tasks]
    return dict(
        tasks=tasks_with_slip, username=session.get('username',''),
        stats={'total': len(all_tasks), 'pending': sum(1 for t in all_tasks if t['status']=='pending'),
               'inprogress': sum(1 for t in all_tasks if t['status']=='inprogress'),
               'completed': sum(1 for t in all_tasks if t['status']=='completed')},
        analytics=build_analytics(all_tasks),
        stamps=stamps,
        stamp_stats={'total_customers': len(stamps), 'total_stamps': sum(v['stamps'] for v in stamps.values()),
                     'total_redeemed': sum(v['rewards_redeemed'] for v in stamps.values())},
        tickets=tickets,
        ticket_stats={'total': len(tickets), 'active': sum(1 for t in tickets.values() if t['status']=='active'),
                      'checked_in': sum(1 for t in tickets.values() if t['status']=='checked_in')},
        stamps_to_reward=STAMPS_TO_REWARD, promptpay_phone=PROMPTPAY_PHONE,
        calendar=cal, ya=yearly_analytics(all_tasks, yr, cal),
        queue_tasks=get_queue_with_tasks(all_tasks, cal),
        qr_image=get_qr_image(), all_slips=slips,
        pending_slips=pending_slips_count(), todos=read_todos(),
        action_items=build_action_items(all_tasks, slips),
        revenue=revenue_analytics(slips),
        status_flow=STATUS_FLOW, status_labels=STATUS_LABELS,
        events=read_events(), notifications=read_notifications()[:20],
        gallery=read_gallery(), reviews=read_reviews(), coupons=read_coupons(),
        crm_customers=crm_summary(all_tasks),
    )

def admin_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if 'username' not in session: return redirect(url_for('login'))
        return f(*a, **kw)
    return dec

# ══════════════════════════════════════════════════════════════════════════════
#  PUBLIC ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/')
def index():
    return render_template('order_form.html', active_page='order')

@app.route('/submit_order', methods=['POST'])
def submit_order():
    saved_files = []
    for f in request.files.getlist('reference_files'):
        if f and f.filename:
            ext = f.filename.rsplit('.', 1)[-1].lower()
            if ext in ALLOWED_ORDER_FILES:
                fname = f"{int(datetime.now().timestamp()*1000)}_{uuid.uuid4().hex[:6]}.{ext}"
                f.save(os.path.join(CUSTOM_ORDER_FOLDER, fname))
                saved_files.append({'filename': fname, 'original': f.filename, 'ext': ext, 'field': 'reference_files'})

    specs_custom = {
        'service_type': request.form.get('service_type', 'custom'),
        'complexity': request.form.get('complexity', 'standard'),
        'quantity': request.form.get('quantity', '1'),
        'width_mm': request.form.get('width_mm', ''),
        'height_mm': request.form.get('height_mm', ''),
        'depth_mm': request.form.get('depth_mm', ''),
        'material': request.form.get('material', ''),
        'finish_level': request.form.get('finish_level', 'basic'),
        'budget': request.form.get('budget', ''),
        'delivery_method': request.form.get('delivery_method', 'pickup'),
        'deadline': request.form.get('deadline',''),
        'reference_files': saved_files,
    }
    spec_lines = [
        f"ประเภทงาน: {specs_custom['service_type']} | ความซับซ้อน: {specs_custom['complexity']}",
        f"จำนวน: {specs_custom['quantity']} | ผิวงาน: {specs_custom['finish_level']}",
    ]
    if specs_custom['width_mm'] or specs_custom['height_mm'] or specs_custom['depth_mm']:
        spec_lines.append(f"ขนาด: {specs_custom['width_mm']}x{specs_custom['height_mm']}x{specs_custom['depth_mm']} mm")
    if specs_custom['material']:
        spec_lines.append(f"วัสดุ/รายละเอียดวัสดุ: {specs_custom['material']}")
    if specs_custom['budget']:
        spec_lines.append(f"งบประมาณ: {specs_custom['budget']}")
    if specs_custom['delivery_method']:
        spec_lines.append(f"รับงาน: {specs_custom['delivery_method']}")
    extra_desc = request.form.get('task_description','').strip()
    if extra_desc:
        spec_lines.extend(['', extra_desc])
    if saved_files:
        spec_lines.append('')
        spec_lines.extend([f"ไฟล์อ้างอิง: {sf['original']}" for sf in saved_files])

    tasks = read_tasks()
    task = {
        'id': str(int(datetime.now().timestamp() * 1000)), 'sn': next_sn(),
        'customer': {'name': request.form.get('customer_name',''), 'phone': request.form.get('customer_phone',''), 'email': request.form.get('customer_email','')},
        'title': request.form.get('task_title',''), 'description': "\n".join(spec_lines),
        'priority': request.form.get('priority','medium'), 'deadline': request.form.get('deadline',''),
        'status': 'pending', 'createdBy': 'ลูกค้า',
        'specs_custom': specs_custom,
        'createdAt': datetime.now().isoformat(), 'updatedAt': datetime.now().isoformat(),
    }
    task['quote'] = {
        'status': 'needed',
        'auto_pricing': calculate_custom_order_price(specs_custom),
        'created_at': _now(),
    }
    tasks.insert(0, task); write_tasks(tasks); code = create_ticket(task)
    add_event(task['id'], 'order_created', 'Order submitted', actor=task['customer']['name'])
    add_event(task['id'], 'quote_needed', 'Waiting for admin quote')
    send_line_admin_notification(task, f"New custom order needs quote: {task.get('sn','')} {task.get('title','')}")
    return render_template('order_form.html', success=True, ticket_code=code,
                           customer_name=task['customer']['name'], task_id=task['id'],
                           order_sn=task.get('sn',''), active_page='order')

# ── FIX 1: /model route — was creating empty tasks on every GET ───────────────
@app.route('/model')
def model():
    """3D order form — just renders the template, does NOT create tasks."""
    return render_template('model.html', active_page='model')  # FIX 1

@app.route('/model/submit', methods=['POST'])
def model_submit():
    """3D printing order - full spec + file upload."""
    saved_files = []
    for fkey in ['model_file', 'ref_image']:
        f = request.files.get(fkey)
        if f and f.filename:
            ext = f.filename.rsplit('.', 1)[-1].lower()
            allowed = ALLOWED_3D if fkey == 'model_file' else ALLOWED_IMG
            if ext in allowed:
                fname = str(int(datetime.now().timestamp()*1000)) + '_' + fkey + '.' + ext
                f.save(os.path.join(MODEL_3D_FOLDER, fname))
                saved_files.append({'field': fkey, 'filename': fname, 'original': f.filename, 'ext': ext})

    specs_3d = {
        'material': request.form.get('material', 'PLA'),
        'color':    request.form.get('color', ''),
        'quality':  request.form.get('quality', 'standard'),
        'infill':   request.form.get('infill', '20'),
        'finish':   request.form.get('finish', 'as_printed'),
        'support':  request.form.get('support', 'auto'),
        'quantity': request.form.get('quantity', '1'),
        'size_x':   request.form.get('size_x', ''),
        'size_y':   request.form.get('size_y', ''),
        'size_z':   request.form.get('size_z', ''),
        'scale':    request.form.get('scale', '100'),
        'use_case': request.form.get('use_case', ''),
        'budget':   request.form.get('budget', ''),
        'deadline': request.form.get('deadline', ''),
        'files':    saved_files,
    }
    for sf in saved_files:
        if sf.get('ext') == 'stl':
            vol = stl_volume_cm3(os.path.join(MODEL_3D_FOLDER, sf['filename']))
            if vol:
                specs_3d['volume_cm3'] = vol
                break

    qty = specs_3d['quantity']
    mat = specs_3d['material']
    q_map = {'draft':'Draft 0.3mm','standard':'Standard 0.2mm','fine':'Fine 0.1mm','ultra':'Ultra 0.05mm'}
    q = q_map.get(specs_3d['quality'], specs_3d['quality'])
    color_val = specs_3d.get('color') or 'ไม่ระบุ'

    lines = [
        "วัสดุ: " + mat + " | สี: " + color_val,
        "คุณภาพ: " + q + " | Infill: " + specs_3d['infill'] + "%",
        "ผิว: " + specs_3d['finish'] + " | Support: " + specs_3d['support'],
        "จำนวน: " + qty + " ชิ้น",
    ]
    if specs_3d['size_x'] or specs_3d['size_y'] or specs_3d['size_z']:
        lines.append("ขนาด: " + specs_3d['size_x'] + "x" + specs_3d['size_y'] + "x" + specs_3d['size_z'] + " mm")
    if specs_3d['scale'] and specs_3d['scale'] != '100':
        lines[-1] = lines[-1] + " (scale " + specs_3d['scale'] + "%)"
    if specs_3d['use_case']:
        lines.append("วัตถุประสงค์: " + specs_3d['use_case'])
    if specs_3d['budget']:
        lines.append("งบประมาณ: " + specs_3d['budget'])
    extra_desc = request.form.get('task_description', '')
    if extra_desc:
        lines.append("")
        lines.append("รายละเอียดเพิ่มเติม:")
        lines.append(extra_desc)
    if saved_files:
        lines.append("")
        for sf in saved_files:
            lines.append("📎 " + sf['field'] + ": " + sf['original'])
    auto_desc = "\n".join(lines)

    tasks = read_tasks()
    title = request.form.get('task_title', '')
    task = {
        'id':       str(int(datetime.now().timestamp() * 1000)),
        'sn':       next_sn(),
        'customer': {
            'name':  request.form.get('customer_name', ''),
            'phone': request.form.get('customer_phone', ''),
            'email': request.form.get('customer_email', ''),
        },
        'title':       "[3D] " + title + " (" + mat + ", " + qty + " ชิ้น)",
        'description': auto_desc,
        'priority':    request.form.get('priority', 'medium'),
        'deadline':    request.form.get('deadline', ''),
        'status':      'pending',
        'createdBy':   'ลูกค้า (3D)',
        'specs_3d':    specs_3d,
        'createdAt':   datetime.now().isoformat(),
        'updatedAt':   datetime.now().isoformat(),
    }
    task['quote'] = {
        'status': 'needed',
        'auto_pricing': calculate_3d_price(specs_3d),
        'created_at': datetime.now().isoformat(),
    }
    tasks.insert(0, task)
    write_tasks(tasks)
    code = create_ticket(task)
    add_event(task['id'], 'order_created', '3D order submitted', actor=task['customer']['name'])
    add_event(task['id'], 'quote_needed', 'Waiting for admin quote')
    send_line_admin_notification(task, f"New 3D order needs quote: {task.get('sn','')} {task.get('title','')}")
    return render_template('model.html', success=True, ticket_code=code,
                           customer_name=task['customer']['name'],
                           task_id=task['id'], order_sn=task.get('sn', ''),
                           saved_files=saved_files, active_page='model')

@app.route('/tracking')
def tracking():
    q = request.args.get('q','').strip().lower()
    if q:
        tasks = read_tasks(); r = tasks
        r = [t for t in r if q in t['customer']['name'].lower() or q in t['customer']['phone'].lower()
             or q in (t.get('sn') or '').lower() or q in t['title'].lower()]
        return render_template('tracking.html', results=r, searched=True, search_q=q, active_page='tracking')
    return render_template('tracking.html', searched=False, active_page='tracking')

@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)

# ── FIX 4: payment() — pass slips= (full list) instead of slip= (single) ──────
@app.route('/payment/<task_id>')
def payment(task_id):
    tasks = read_tasks(); task = next((t for t in tasks if t['id'] == task_id), None)
    if not task: return 'ไม่พบออเดอร์', 404
    tickets = read_tickets(); code = next((c for c,tk in tickets.items() if tk['task_id']==task_id), '')
    amount = request.args.get('amount', type=float)
    pay_part = request.args.get('part', '')
    quote = task.get('quote', {})
    if amount is None and quote.get('status') == 'approved':
        if pay_part == 'balance':
            amount = quote.get('balance_amount') or quote.get('amount')
        else:
            amount = quote.get('deposit_amount') or quote.get('amount')
    qr_image = get_qr_image()
    payload = generate_promptpay_payload(PROMPTPAY_PHONE, amount) if not qr_image else ''
    return render_template('payment.html', task=task, ticket_code=code,
                           promptpay_payload=payload, promptpay_phone=PROMPTPAY_PHONE,
                           amount=amount, qr_image=qr_image,
                           slips=slips_for_task(task_id))  # FIX 4: slips= list

# ── FIX 3: upload_slip — read 'slip' field not 'slip_file' ────────────────────
@app.route('/upload_slip/<task_id>', methods=['POST'])
def upload_slip(task_id):
    tasks = read_tasks(); task = next((t for t in tasks if t['id']==task_id), None)
    if not task: return 'ไม่พบออเดอร์', 404
    file = request.files.get('slip')   # FIX 3: was 'slip_file'
    if not file or file.filename == '': return redirect(url_for('payment', task_id=task_id)+'?error=no_file')
    if not allowed_file(file.filename): return redirect(url_for('payment', task_id=task_id)+'?error=bad_type')
    ext = file.filename.rsplit('.',1)[1].lower()
    fname = f"{task_id}_{int(datetime.now().timestamp())}.{ext}"
    file.save(os.path.join(SLIP_FOLDER, fname))
    slips = read_slips()
    slips.setdefault(task_id, []).append({'file': f'slips/{fname}', 'uploaded_at': datetime.now().isoformat(),
                                           'status': 'pending', 'note': '', 'amount': request.form.get('amount','')})
    write_slips(slips)
    add_event(task_id, 'slip_uploaded', f"Amount: {request.form.get('amount','')}", actor=task['customer'].get('name','customer'))
    send_line_admin_notification(task, f"Payment slip uploaded for {task.get('sn','')} amount {request.form.get('amount','')}")
    tickets = read_tickets(); code = next((c for c,tk in tickets.items() if tk['task_id']==task_id), '')
    return render_template('payment.html', task=task, ticket_code=code,
                           promptpay_payload='', promptpay_phone=PROMPTPAY_PHONE,
                           amount=None, qr_image=get_qr_image(),
                           slips=slips_for_task(task_id),  # FIX 4
                           slip_uploaded=True)

@app.route('/order_pdf/<task_id>')
def order_pdf(task_id):
    if not PDF_ENABLED: return 'PDF ไม่พร้อมใช้งาน (ติดตั้ง reportlab)', 500
    try:
        tasks = read_tasks(); task = next((t for t in tasks if t['id']==task_id), None)
        if not task: return 'ไม่พบออเดอร์', 404
        tickets = read_tickets(); code = next((c for c,tk in tickets.items() if tk['task_id']==task_id), '')
        amount = request.args.get('amount', type=float)
        qr_img = get_qr_image(); payload = generate_promptpay_payload(PROMPTPAY_PHONE, amount) if (amount and not qr_img) else ''
        pdf_bytes = generate_order_pdf(task, code, payload, COMPANY_NAME)
        resp = make_response(pdf_bytes)
        resp.headers['Content-Type'] = 'application/pdf'
        resp.headers['Content-Disposition'] = f'attachment; filename="order_{task_id[-6:]}.pdf"'
        return resp
    except Exception as e:
        return f'เกิดข้อผิดพลาด: {e}', 500

@app.route('/admin/order_pdf/<task_id>')
@admin_required
def admin_order_pdf(task_id):
    if not PDF_ENABLED: return 'PDF ไม่พร้อมใช้งาน', 500
    tasks = read_tasks(); task = next((t for t in tasks if t['id']==task_id), None)
    if not task: return 'ไม่พบ', 404
    tickets = read_tickets(); code = next((c for c,tk in tickets.items() if tk['task_id']==task_id), '')
    inc_qr = request.args.get('qr','0') == '1'; amount = request.args.get('amount', type=float)
    qr_img = get_qr_image(); payload = generate_promptpay_payload(PROMPTPAY_PHONE, amount) if (inc_qr and not qr_img) else ''
    pdf_bytes = generate_order_pdf(task, code, payload, COMPANY_NAME)
    resp = make_response(pdf_bytes)
    resp.headers['Content-Type'] = 'application/pdf'
    resp.headers['Content-Disposition'] = f'attachment; filename="order_{task_id[-6:]}.pdf"'
    return resp

@app.route('/admin/invoice/<task_id>')
@admin_required
def admin_invoice(task_id):
    if not PDF_ENABLED: return 'PDF not ready', 500
    tasks = read_tasks(); task = find_task(tasks, task_id)
    if not task: return 'not found', 404
    inv = invoice_for_task(task)
    inv_task = {**task, 'title': f"Invoice {inv['invoice_no']} - {task.get('title','')}",
                'description': f"Subtotal: {inv['subtotal']:.2f}\nVAT {inv['vat_rate']:.1f}%: {inv['vat']:.2f}\nTotal: {inv['total']:.2f}"}
    pdf_bytes = generate_order_pdf(inv_task, '', '', COMPANY_NAME)
    resp = make_response(pdf_bytes)
    resp.headers['Content-Type'] = 'application/pdf'
    resp.headers['Content-Disposition'] = f'attachment; filename="{inv["invoice_no"]}.pdf"'
    if request.args.get('email') == '1':
        send_email_notification(task, f'Invoice {inv["invoice_no"]}', 'Your invoice is ready. Please see the attached PDF in the admin download for now.')
    return resp

@app.route('/ticket/<code>')
def view_ticket(code):
    tickets = read_tickets(); ticket = tickets.get(code.upper())
    if not ticket: return render_template('ticket.html', error=True, code=code, ticket=None, task=None)
    tasks = read_tasks(); task = next((t for t in tasks if t['id']==ticket['task_id']), None)
    return render_template('ticket.html', ticket=ticket, code=code.upper(), task=task)

@app.route('/checkin', methods=['GET','POST'])
def public_checkin():
    msg = None; ticket = None; code = ''
    if request.method == 'POST':
        code = request.form.get('code','').strip().upper(); tickets = read_tickets()
        if code not in tickets: msg = ('error', f'ไม่พบ Ticket รหัส {code}')
        elif tickets[code]['status'] == 'checked_in':
            msg = ('warning', f'Check-in แล้วเมื่อ {tickets[code]["checked_in_at"][:16].replace("T"," ")}'); ticket = tickets[code]
        else:
            tickets[code].update({'status':'checked_in','checked_in_at':datetime.now().isoformat(),'checked_in_by':'self'})
            write_tickets(tickets); ticket = tickets[code]
            msg = ('success', f'Check-in สำเร็จ! ยินดีต้อนรับ {ticket["customer_name"]} 🎉')
    return render_template('checkin.html', message=msg, ticket=ticket, code=code, active_page='checkin')

@app.route('/webhook', methods=['GET','POST'])
def line_webhook():
    if request.method == 'GET': return jsonify({'status': 'LINE Webhook active'})
    if not LINE_ENABLED: return jsonify({'error': 'LINE not configured'}), 500
    body = request.get_data(); sig = request.headers.get('X-Line-Signature','')
    if not verify_signature(body, sig): return jsonify({'error': 'Bad signature'}), 403
    try: handle_events(json.loads(body).get('events',[]), read_tasks, read_tickets)
    except Exception as e: app.logger.error(f'LINE: {e}')
    return jsonify({'status': 'ok'})

@app.route('/login', methods=['GET','POST'])
def login():
    users = read_users(); ft = len(users) == 0
    if request.method == 'POST':
        ip = request.headers.get('X-Forwarded-For', request.remote_addr or 'local').split(',')[0].strip()
        attempts = [t for t in _LOGIN_ATTEMPTS.get(ip, []) if time.time() - t < 900]
        if len(attempts) >= 5:
            return render_template('login.html', error='พยายามเข้าสู่ระบบมากเกินไป กรุณารอ 15 นาที', first_time=ft, active_page='login'), 429
        u = request.form.get('username',''); p = request.form.get('password','')
        if ft:
            users[u] = hash_password(p); write_users(users); session['username'] = u; _LOGIN_ATTEMPTS.pop(ip, None); return redirect(url_for('admin_dashboard'))
        elif u in users and verify_password(users[u], p):
            if password_needs_upgrade(users[u]):
                users[u] = hash_password(p); write_users(users)
            session['username'] = u; _LOGIN_ATTEMPTS.pop(ip, None); return redirect(url_for('admin_dashboard'))
        else:
            attempts.append(time.time()); _LOGIN_ATTEMPTS[ip] = attempts
            return render_template('login.html', error='ชื่อผู้ใช้หรือรหัสผ่านไม่ถูกต้อง', first_time=ft, active_page='login')
    return render_template('login.html', first_time=ft, active_page='login')

@app.route('/logout')
def logout():
    session.pop('username', None); return redirect(url_for('index'))

@app.route('/contact', methods=['GET','POST'])
def contact():
    sent = False
    if request.method == 'POST':
        name = request.form.get('name','').strip(); email = request.form.get('email','').strip()
        phone = request.form.get('phone','').strip()
        if name and (email or phone): sent = True
    return render_template('contact.html', sent=sent, promptpay_phone=PROMPTPAY_PHONE, active_page='contact')

@app.route('/gallery')
def gallery():
    items = [g for g in read_gallery() if g.get('active', True)]
    return render_template('gallery.html', items=items, reviews=read_reviews(), active_page='gallery')

@app.route('/review/<task_id>', methods=['GET','POST'])
def review_order(task_id):
    tasks = read_tasks(); task = find_task(tasks, task_id)
    if not task: return 'not found', 404
    if request.method == 'POST':
        reviews = read_reviews()
        row = {
            'id': uuid.uuid4().hex[:12], 'task_id': task_id,
            'name': request.form.get('name') or task.get('customer', {}).get('name', ''),
            'rating': int(request.form.get('rating') or 5),
            'comment': request.form.get('comment','').strip(),
            'active': True, 'created_at': _now(),
        }
        reviews.insert(0, row); write_reviews(reviews)
        add_event(task_id, 'review_submitted', f"{row['rating']} stars", actor=row['name'])
        return render_template('review.html', task=task, submitted=True, review=row)
    return render_template('review.html', task=task, submitted=False)

# ══════════════════════════════════════════════════════════════════════════════
#  ADMIN ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/admin')
@admin_required
def admin_dashboard():
    return render_template('admin_dashboard.html', products_count=len(read_products()), **admin_context())

@app.route('/admin/filter/<status>')
@admin_required
def filter_tasks(status):
    all_tasks = read_tasks(); tasks = all_tasks if status=='all' else [t for t in all_tasks if t['status']==status]
    ctx = admin_context(tasks); ctx['current_filter'] = status
    return render_template('admin_dashboard.html', **ctx)

@app.route('/admin/update_status', methods=['POST'])
@admin_required
def update_status():
    task_id = request.form.get('task_id','').strip()
    new_status = request.form.get('new_status','').strip()
    if new_status not in set(STATUS_FLOW):
        return jsonify({'status':'error','error':'invalid status'}), 400
    tasks = read_tasks()
    for t in tasks:
        if t['id'] == task_id:
            old_status = t.get('status', 'pending')
            if new_status == 'completed' and t['status'] != 'completed':
                add_stamp(t['customer']['phone'], t['customer']['name'])
            t['status'] = new_status; t['updatedAt'] = datetime.now().isoformat()
            t['updatedBy'] = session.get('username','')
            if old_status != new_status:
                add_event(task_id, 'status_changed', f"{STATUS_LABELS.get(old_status, old_status)} → {STATUS_LABELS.get(new_status, new_status)}")
                notify_status_change(t, old_status, new_status)
            break
    write_tasks(tasks)
    updated = next((t for t in tasks if t['id']==task_id), None)
    return jsonify({'status':'ok','task':updated})

@app.route('/admin/delete', methods=['POST'])
@admin_required
def delete_task():
    tid = request.form.get('task_id')
    write_tasks([t for t in read_tasks() if t['id']!=tid])
    add_event(tid, 'task_deleted', 'Task removed from admin dashboard')
    return jsonify({'status':'ok'})

@app.route('/admin/task_events/<task_id>')
@admin_required
def admin_task_events(task_id):
    return jsonify({'status': 'ok', 'events': events_for_task(task_id)})

@app.route('/admin/notifications')
@admin_required
def admin_notifications():
    rows = read_notifications()[:20]
    events = read_events()[:20]
    return jsonify({'status':'ok','notifications':rows,'events':events,'count':len(rows)+len(events)})

@app.route('/admin/quote', methods=['POST'])
@admin_required
def admin_quote():
    task_id = request.form.get('task_id','')
    tasks = read_tasks(); task = find_task(tasks, task_id)
    if not task: return jsonify({'status':'error','error':'not found'}), 404
    overrides = {
        'discount': request.form.get('discount', 0) or 0,
        'deposit_percent': request.form.get('deposit_percent', 50) or 50,
        'minimum': request.form.get('minimum', 150) or 150,
    }
    if task.get('specs_3d'):
        auto = calculate_3d_price(task.get('specs_3d', {}), overrides)
    else:
        auto = calculate_custom_order_price(task.get('specs_custom', {}), overrides)
    amount = float(request.form.get('amount') or auto['amount'])
    deposit_percent = float(request.form.get('deposit_percent') or auto['deposit_percent'])
    deposit_amount = round(amount * deposit_percent / 100, 2)
    task['quote'] = {
        **auto, 'amount': round(amount, 2), 'deposit_percent': deposit_percent,
        'deposit_amount': deposit_amount, 'balance_amount': round(amount - deposit_amount, 2),
        'note': request.form.get('note','').strip(), 'status': 'sent',
        'sent_at': _now(), 'sent_by': session.get('username',''),
    }
    task['status'] = 'quoted'
    task['updatedAt'] = _now()
    write_tasks(tasks)
    add_event(task_id, 'quote_sent', f"Quote {amount:.2f} THB, deposit {deposit_amount:.2f} THB")
    send_email_notification(task, 'Quote ready for your order', f"Quote amount: {amount:.2f} THB\nDeposit: {deposit_amount:.2f} THB")
    return jsonify({'status':'ok','quote':task['quote'], 'task': task})

@app.route('/task/comment', methods=['POST'])
def task_comment():
    task_id = request.form.get('task_id','')
    text = request.form.get('comment','').strip()
    if not text: return jsonify({'status':'error','error':'empty'}), 400
    tasks = read_tasks(); task = find_task(tasks, task_id)
    if not task: return jsonify({'status':'error','error':'not found'}), 404
    is_admin = bool(session.get('username'))
    is_customer = session.get('customer_phone') == task.get('customer', {}).get('phone')
    if not (is_admin or is_customer):
        return jsonify({'status':'error','error':'forbidden'}), 403
    ev = add_event(task_id, 'comment_admin' if is_admin else 'comment_customer', text)
    if is_customer:
        send_line_admin_notification(task, f"Customer comment on {task.get('sn','')}: {text[:120]}")
    return jsonify({'status':'ok','event':ev})

@app.route('/admin/bulk/status', methods=['POST'])
@admin_required
def admin_bulk_status():
    ids = request.form.getlist('task_ids') or [x for x in request.form.get('task_ids','').split(',') if x]
    new_status = request.form.get('new_status','')
    if new_status not in set(STATUS_FLOW): return jsonify({'status':'error','error':'invalid status'}), 400
    tasks = read_tasks(); changed = []
    for t in tasks:
        if t.get('id') in ids:
            old = t.get('status','pending')
            t['status'] = new_status; t['updatedAt'] = _now(); t['updatedBy'] = session.get('username','')
            changed.append(t['id']); add_event(t['id'], 'bulk_status_changed', f"{old} → {new_status}")
    write_tasks(tasks)
    return jsonify({'status':'ok','changed':changed})

@app.route('/quote/<task_id>/<action>', methods=['POST'])
def customer_quote_action(task_id, action):
    if action not in {'approve', 'reject'}:
        return 'invalid action', 400
    tasks = read_tasks(); task = find_task(tasks, task_id)
    if not task: return 'not found', 404
    quote = task.setdefault('quote', {})
    quote['status'] = 'approved' if action == 'approve' else 'rejected'
    quote['responded_at'] = _now()
    quote['customer_note'] = request.form.get('note','')
    if action == 'approve':
        task['status'] = 'approved'
    task['updatedAt'] = _now()
    write_tasks(tasks)
    add_event(task_id, 'quote_' + quote['status'], quote.get('customer_note',''), actor=task.get('customer', {}).get('name','customer'))
    send_line_admin_notification(task, f"Quote {quote['status']} for {task.get('sn','')}")
    if action == 'approve':
        return redirect(url_for('payment', task_id=task_id))
    return redirect(url_for('customer_dashboard'))

@app.route('/admin/backup')
@admin_required
def admin_backup():
    path = create_backup_zip(auto=False)
    with open(path, 'rb') as f: data = f.read()
    resp = make_response(data)
    resp.headers['Content-Type'] = 'application/zip'
    resp.headers['Content-Disposition'] = f'attachment; filename="zerphyrus_backup_{datetime.now().strftime("%Y%m%d_%H%M")}.zip"'
    return resp

@app.route('/admin/restore', methods=['POST'])
@admin_required
def admin_restore():
    file = request.files.get('backup')
    if not file or not file.filename.endswith('.zip'):
        return jsonify({'status':'error','error':'zip required'}), 400
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
    file.save(tmp.name); tmp.close()
    restored = []
    try:
        with zipfile.ZipFile(tmp.name) as z:
            for info in z.infolist():
                name = info.filename.replace('\\','/')
                if '..' in name or name.startswith('/'):
                    continue
                if name.endswith('.json') and '/' not in name:
                    with z.open(info) as src, open(name, 'wb') as dst:
                        shutil.copyfileobj(src, dst)
                    restored.append(name)
                elif name.startswith('uploads/'):
                    target = os.path.join(PROJECT_DIR, name)
                    os.makedirs(os.path.dirname(target), exist_ok=True)
                    with z.open(info) as src, open(target, 'wb') as dst:
                        shutil.copyfileobj(src, dst)
                    restored.append(name)
        add_event('system', 'backup_restored', ', '.join(restored[:20]))
        return jsonify({'status':'ok','restored':restored})
    finally:
        try: os.unlink(tmp.name)
        except Exception: pass

@app.route('/admin/migrate_sqlite', methods=['POST'])
@admin_required
def admin_migrate_sqlite():
    db_path = migrate_json_to_sqlite()
    return jsonify({'status':'ok','db':db_path})

@app.route('/admin/coupons/add', methods=['POST'])
@admin_required
def admin_coupon_add():
    coupons = read_coupons()
    code = request.form.get('code','').strip().upper()
    if not code: return jsonify({'status':'error','error':'code required'}), 400
    coupons = [c for c in coupons if c.get('code') != code]
    coupons.insert(0, {
        'code': code, 'type': request.form.get('type','fixed'),
        'value': float(request.form.get('value') or 0),
        'active': True, 'created_at': _now(), 'created_by': session.get('username',''),
    })
    write_coupons(coupons)
    return jsonify({'status':'ok','coupons':coupons})

@app.route('/admin/customer/<phone>')
@admin_required
def admin_customer_detail(phone):
    customers = read_customers(); customer = customers.get(phone, {'phone': phone})
    tasks = [t for t in read_tasks() if t.get('customer', {}).get('phone') == phone]
    total_spend = sum(float(t.get('quote', {}).get('amount') or t.get('order_total') or 0) for t in tasks)
    return jsonify({'status':'ok','customer':customer,'orders':tasks,'order_count':len(tasks),'total_spend':round(total_spend,2)})

@app.route('/admin/customer/<phone>/tags', methods=['POST'])
@admin_required
def admin_customer_tags(phone):
    customers = read_customers(); c = customers.setdefault(phone, {'phone': phone})
    c['tags'] = [x.strip() for x in request.form.get('tags','').split(',') if x.strip()]
    c['note'] = request.form.get('note','')
    write_customers(customers)
    return jsonify({'status':'ok','customer':c})

@app.route('/admin/gallery/add', methods=['POST'])
@admin_required
def admin_gallery_add():
    task_id = request.form.get('task_id','')
    tasks = read_tasks(); task = find_task(tasks, task_id) if task_id else None
    file = request.files.get('image')
    image_path = request.form.get('image','').strip()
    if file and file.filename and allowed_file(file.filename):
        ext = file.filename.rsplit('.', 1)[1].lower()
        fname = f"{int(datetime.now().timestamp())}_{secure_filename(file.filename)}"
        if not fname.lower().endswith('.' + ext):
            fname += '.' + ext
        file.save(os.path.join(GALLERY_FOLDER, fname))
        image_path = f'gallery/{fname}'
    if not image_path:
        return jsonify({'status':'error','error':'image required'}), 400
    gallery_items = read_gallery()
    item = {
        'id': uuid.uuid4().hex[:12], 'task_id': task_id,
        'title': request.form.get('title') or (task.get('title') if task else 'Finished work'),
        'description': request.form.get('description',''),
        'image': image_path, 'active': True, 'created_at': _now(),
    }
    gallery_items.insert(0, item); write_gallery(gallery_items)
    if task_id: add_event(task_id, 'gallery_added', item['title'])
    return jsonify({'status':'ok','item':item})

@app.route('/admin/job_sheet/<task_id>')
@admin_required
def admin_job_sheet(task_id):
    tasks = read_tasks(); task = find_task(tasks, task_id)
    if not task: return 'not found', 404
    return render_template('job_sheet.html', task=task, events=events_for_task(task_id), status_labels=STATUS_LABELS)

@app.route('/admin/upload_qr', methods=['POST'])
@admin_required
def upload_qr():
    file = request.files.get('qr')
    wants_json = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if not file or not allowed_file(file.filename):
        if wants_json:
            return jsonify({'status': 'error', 'message': 'bad_file'}), 400
        return redirect(url_for('line_config')+'?error=bad_file')
    for ext in ALLOWED_IMG:
        old = os.path.join(QR_FOLDER, f'promptpay.{ext}')
        if os.path.exists(old): os.remove(old)
    ext = file.filename.rsplit('.',1)[1].lower()
    file.save(os.path.join(QR_FOLDER, f'promptpay.{ext}'))
    if wants_json:
        return jsonify({'status': 'ok', 'qr_image': f'qr/promptpay.{ext}'})
    return redirect(url_for('line_config')+'?qr_saved=1')

@app.route('/admin/delete_qr', methods=['POST'])
@admin_required
def delete_qr():
    for ext in ALLOWED_IMG:
        p = os.path.join(QR_FOLDER, f'promptpay.{ext}')
        if os.path.exists(p): os.remove(p)
    return redirect(url_for('line_config'))

@app.route('/admin/verify_slip', methods=['POST'])
@admin_required
def verify_slip():
    task_id = request.form.get('task_id',''); slip_idx = int(request.form.get('slip_idx',0))
    action = request.form.get('action','approve'); note = request.form.get('note','')
    slips = read_slips(); task_slips = slips.get(task_id, [])
    if 0 <= slip_idx < len(task_slips):
        task_slips[slip_idx]['status'] = 'approved' if action=='approve' else 'rejected'
        task_slips[slip_idx]['note'] = note; task_slips[slip_idx]['verified_at'] = datetime.now().isoformat()
        task_slips[slip_idx]['verified_by'] = session.get('username','')
        add_event(task_id, 'slip_verified', task_slips[slip_idx]['status'] + (f": {note}" if note else ''))
        if action == 'approve':
            tasks = read_tasks()
            for t in tasks:
                if t['id']==task_id and t['status']=='pending':
                    t['status']='inprogress'; t['updatedAt']=datetime.now().isoformat()
                    add_event(task_id, 'status_changed', 'Pending → In Progress')
                    notify_status_change(t, 'pending', 'inprogress')
            write_tasks(tasks)
    write_slips(slips)
    return jsonify({'status':'ok','slip_status': task_slips[slip_idx]['status'] if 0 <= slip_idx < len(task_slips) else ''})

@app.route('/admin/redeem_stamp', methods=['POST'])
@admin_required
def redeem_stamp():
    phone = request.form.get('phone',''); s = read_stamps()
    if phone in s and s[phone]['stamps'] >= STAMPS_TO_REWARD:
        s[phone]['stamps'] -= STAMPS_TO_REWARD; s[phone]['rewards_redeemed'] += 1; write_stamps(s)
    s = read_stamps()
    return jsonify({'status':'ok','stamps': s.get(phone,{}).get('stamps',0), 'redeemed': s.get(phone,{}).get('rewards_redeemed',0)})

@app.route('/admin/add_stamp_manual', methods=['POST'])
@admin_required
def add_stamp_manual():
    phone = request.form.get('phone','').strip(); name = request.form.get('name','').strip()
    if phone: add_stamp(phone, name)
    s = read_stamps()
    return jsonify({'status':'ok','phone':phone,'stamps': s.get(phone,{}).get('stamps',0)})

@app.route('/admin/checkin_ticket', methods=['POST'])
@admin_required
def admin_checkin_ticket():
    code = request.form.get('code','').strip().upper(); tickets = read_tickets()
    if code in tickets and tickets[code]['status']=='active':
        tickets[code].update({'status':'checked_in','checked_in_at':datetime.now().isoformat(),'checked_in_by':session.get('username','')})
        write_tickets(tickets)
    return jsonify({'status':'ok','code':code})

@app.route('/admin/queue/reorder', methods=['POST'])
@admin_required
def api_reorder_queue():
    data = request.get_json(silent=True) or {}; reorder_queue(data.get('order',[])); return jsonify({'status':'ok'})

@app.route('/admin/queue/estimate', methods=['POST'])
@admin_required
def api_set_estimate():
    set_task_estimate(request.form.get('task_id',''), float(request.form.get('hours',0) or 0), request.form.get('note',''))
    return jsonify({'status':'ok'})

@app.route('/admin/calendar/settings', methods=['POST'])
@admin_required
def api_calendar_settings():
    work_days = [int(d) for d in request.form.getlist('work_days')]
    update_calendar_settings(work_days, int(request.form.get('capacity',3)))
    return jsonify({'status':'ok'})

@app.route('/admin/calendar/add_date', methods=['POST'])
@admin_required
def api_add_date():
    ds = request.form.get('date',''); dtype = request.form.get('type','holiday'); note = request.form.get('note','')
    if ds: add_custom_date(ds, dtype, note)
    return jsonify({'status':'ok'})

@app.route('/admin/calendar/remove_date', methods=['POST'])
@admin_required
def api_remove_date():
    ds = request.form.get('date','')
    if ds: remove_custom_date(ds)
    return jsonify({'status':'ok'})

@app.route('/admin/api/yearly/<int:year>')
@admin_required
def api_yearly(year):
    return jsonify(yearly_analytics(read_tasks(), year, read_calendar()))

@app.route('/admin/line_config', methods=['GET','POST'])
@admin_required
def line_config():
    msg = None; qr_saved = request.args.get('qr_saved') == '1'
    if request.method == 'POST':
        lines = []
        for k in ['LINE_CHANNEL_ACCESS_TOKEN','LINE_CHANNEL_SECRET','ADMIN_LINE_USER_ID',
                  'PROMPTPAY_PHONE','COMPANY_NAME','PREFERRED_SCHEME',
                  'SMTP_HOST','SMTP_PORT','SMTP_USER','SMTP_PASSWORD','SMTP_FROM','SMTP_TLS']:
            v = request.form.get(k,'').strip()
            if v: os.environ[k] = v; lines.append(f"{k}={v}")
        if lines:
            with open('.env', 'w', encoding='utf-8') as f: f.write('\n'.join(lines)+'\n')
        msg = 'บันทึกเรียบร้อย — รีสตาร์ทเซิร์ฟเวอร์เพื่อให้มีผล'
    return render_template('line_config.html', msg=msg, qr_saved=qr_saved, qr_image=get_qr_image(),
                           token=os.environ.get('LINE_CHANNEL_ACCESS_TOKEN',''),
                           secret=os.environ.get('LINE_CHANNEL_SECRET',''),
                           admin_id=os.environ.get('ADMIN_LINE_USER_ID',''),
                           promptpay=PROMPTPAY_PHONE, company=COMPANY_NAME, scheme=PREFERRED_SCHEME,
                           smtp_host=os.environ.get('SMTP_HOST',''), smtp_port=os.environ.get('SMTP_PORT','587'),
                           smtp_user=os.environ.get('SMTP_USER',''), smtp_password=os.environ.get('SMTP_PASSWORD',''),
                           smtp_from=os.environ.get('SMTP_FROM',''), smtp_tls=os.environ.get('SMTP_TLS','1'),
                           webhook_url=get_webhook_url())

@app.route('/admin/todos/add', methods=['POST'])
@admin_required
def admin_todos_add():
    todos = read_todos()
    todo = {'id': str(int(datetime.now().timestamp()*1000)), 'text': request.form.get('text','').strip(),
            'done': False, 'priority': request.form.get('priority','medium'), 'due': request.form.get('due',''),
            'createdAt': datetime.now().isoformat()}
    if not todo['text']: return jsonify({'status':'error','error':'empty'}), 400
    todos.insert(0, todo); write_todos(todos)
    return jsonify({'status':'ok','todo':todo})

@app.route('/admin/todos/toggle/<tid>', methods=['POST'])
@admin_required
def admin_todos_toggle(tid):
    todos = read_todos()
    for t in todos:
        if t['id']==tid: t['done'] = not t['done']; break
    write_todos(todos); return jsonify({'status':'ok'})

@app.route('/admin/todos/delete/<tid>', methods=['POST'])
@admin_required
def admin_todos_delete(tid):
    write_todos([t for t in read_todos() if t['id']!=tid]); return jsonify({'status':'ok'})

# ── Product catalog ────────────────────────────────────────────────────────────
def _cart_key():
    if 'cart_id' not in session: session['cart_id'] = _secrets.token_hex(16)
    return session['cart_id']

def get_cart():
    cid = _cart_key(); carts = _r('orders_cart.json', {}); return carts.get(cid, [])

def save_cart(items):
    cid = _cart_key(); carts = _r('orders_cart.json', {}); carts[cid] = items; _w('orders_cart.json', carts)

def cart_count(): return sum(i['qty'] for i in get_cart())

@app.route('/catalog')
def catalog():
    products = [p for p in read_products() if p.get('active', True)]
    category = request.args.get('cat',''); search = request.args.get('q','').lower()
    cats = sorted({p.get('category','') for p in products if p.get('category')})
    if category: products = [p for p in products if p.get('category','')==category]
    if search: products = [p for p in products if search in p['name'].lower() or search in p.get('description','').lower()]
    return render_template('catalog.html', products=products, categories=cats, active_cat=category, search=search)

@app.route('/product/<pid>')
def product_detail(pid):
    products = read_products(); product = next((p for p in products if p['id']==pid), None)
    if not product or not product.get('active',True): return redirect(url_for('catalog'))
    return render_template('product_detail.html', product=product)

@app.route('/cart')
def view_cart():
    cart = get_cart(); products = read_products(); pid_map = {p['id']:p for p in products}
    items = []; total = 0
    for item in cart:
        p = pid_map.get(item['product_id'])
        if p:
            subtotal = p['price'] * item['qty']; total += subtotal
            items.append({**item,'product':p,'subtotal':subtotal})
    return render_template('cart.html', items=items, total=total)

@app.route('/cart/add', methods=['POST'])
def cart_add():
    pid = request.form.get('product_id',''); qty = max(1, int(request.form.get('qty',1)))
    back = request.form.get('back', url_for('catalog'))
    products = read_products(); product = next((p for p in products if p['id']==pid and p.get('active',True)), None)
    if not product: return redirect(back)
    cart = get_cart()
    for item in cart:
        if item['product_id']==pid:
            item['qty'] = min(item['qty']+qty, product.get('stock',9999)); break
    else: cart.append({'product_id':pid,'qty':qty})
    save_cart(cart)
    if request.form.get('buy_now'): return redirect(url_for('cart_checkout'))
    return redirect(back)

@app.route('/cart/update', methods=['POST'])
def cart_update():
    pid = request.form.get('product_id',''); qty = int(request.form.get('qty',0))
    products = read_products(); product = next((p for p in products if p['id']==pid), None)
    cart = get_cart()
    if qty <= 0: cart = [i for i in cart if i['product_id']!=pid]
    else:
        stock = product.get('stock') if product else None
        if stock is not None: qty = min(qty, stock)
        for item in cart:
            if item['product_id']==pid: item['qty']=qty; break
    save_cart(cart); return redirect(url_for('view_cart'))

@app.route('/cart/remove', methods=['POST'])
def cart_remove():
    pid = request.form.get('product_id',''); save_cart([i for i in get_cart() if i['product_id']!=pid])
    return redirect(url_for('view_cart'))

@app.route('/cart/checkout', methods=['GET','POST'])
def cart_checkout():
    cart = get_cart()
    if not cart: return redirect(url_for('catalog'))
    products = read_products(); pid_map = {p['id']:p for p in products}
    items = []; total = 0
    for item in cart:
        p = pid_map.get(item['product_id'])
        if p:
            subtotal = p['price']*item['qty']; total += subtotal
            items.append({**item,'product':p,'subtotal':subtotal})
    if request.method == 'POST':
        name = request.form.get('name',''); phone = request.form.get('phone','')
        email = request.form.get('email',''); addr = request.form.get('address','')
        tasks = read_tasks()
        desc = 'สินค้า:\n' + '\n'.join(f'- {i["product"]["name"]} x{i["qty"]}  ฿{i["subtotal"]:.0f}' for i in items) + f'\n\nที่อยู่จัดส่ง: {addr}'
        task = {'id': str(int(datetime.now().timestamp()*1000)), 'sn': next_sn(),
                'customer': {'name':name,'phone':phone,'email':email},
                'title': f'คำสั่งซื้อออนไลน์ ({len(items)} รายการ)', 'description': desc,
                'priority': 'medium', 'deadline': '', 'status': 'pending', 'createdBy': 'ลูกค้า (cart)',
                'createdAt': datetime.now().isoformat(), 'updatedAt': datetime.now().isoformat(),
                'order_total': total}
        tasks.insert(0, task); write_tasks(tasks); create_ticket(task); save_cart([])
        for item in items:
            for p in products:
                if p['id']==item['product_id'] and p.get('stock') is not None:
                    p['stock'] = max(0, p['stock']-item['qty'])
        write_products(products)
        return redirect(url_for('payment', task_id=task['id'])+f'?amount={total:.0f}')
    return render_template('checkout.html', items=items, total=total)

@app.route('/admin/products')
@admin_required
def admin_products():
    return render_template('admin_products.html', products=read_products(), username=session.get('username',''))

@app.route('/admin/products/add', methods=['POST'])
@admin_required
def admin_product_add():
    products = read_products(); pid = str(int(datetime.now().timestamp()*1000))
    stock_val = request.form.get('stock','').strip()
    product = {'id':pid, 'name':request.form.get('name','').strip(),
               'description':request.form.get('description','').strip(),
               'price':float(request.form.get('price',0) or 0),
               'category':request.form.get('category','').strip(),
               'stock':int(stock_val) if stock_val else None,
               'active':request.form.get('active')=='on', 'image':'',
               'createdAt':datetime.now().isoformat()}
    file = request.files.get('image')
    if file and file.filename and allowed_file(file.filename):
        ext = file.filename.rsplit('.',1)[1].lower(); fname = f'{pid}.{ext}'
        _, thumb = save_optimized_upload(file, PRODUCT_IMG_FOLDER, fname)
        product['image'] = f'products/{fname}'; product['thumb'] = f'products/{thumb}' if thumb else product['image']
    products.insert(0, product); write_products(products)
    return jsonify({'status':'ok','id':pid,'name':product['name']})

@app.route('/admin/products/edit/<pid>', methods=['POST'])
@admin_required
def admin_product_edit(pid):
    products = read_products(); updated_name = ''
    for p in products:
        if p['id']==pid:
            p['name'] = request.form.get('name',p['name']).strip()
            p['description'] = request.form.get('description',p.get('description','')).strip()
            try: p['price'] = float(request.form.get('price',p['price']) or 0)
            except: pass
            stock_val = request.form.get('stock','').strip()
            p['stock'] = int(stock_val) if stock_val else None
            p['category'] = request.form.get('category',p.get('category','')).strip()
            p['active'] = request.form.get('active','off') == 'on'
            updated_name = p['name']
            file = request.files.get('image')
            if file and file.filename and allowed_file(file.filename):
                ext = file.filename.rsplit('.',1)[1].lower(); fname = f'{pid}.{ext}'
                _, thumb = save_optimized_upload(file, PRODUCT_IMG_FOLDER, fname)
                p['image'] = f'products/{fname}'; p['thumb'] = f'products/{thumb}' if thumb else p['image']
            break
    write_products(products); return jsonify({'status':'ok','name':updated_name})

@app.route('/admin/products/delete/<pid>', methods=['POST'])
@admin_required
def admin_product_delete(pid):
    write_products([p for p in read_products() if p['id']!=pid])
    for ext in ALLOWED_IMG:
        path = os.path.join(PRODUCT_IMG_FOLDER, f'{pid}.{ext}')
        if os.path.exists(path): os.remove(path)
    return jsonify({'status':'ok'})

@app.route('/admin/products/toggle/<pid>', methods=['POST'])
@admin_required
def admin_product_toggle(pid):
    products = read_products()
    for p in products:
        if p['id']==pid: p['active'] = not p.get('active',True); break
    write_products(products); return jsonify({'status':'ok'})

@app.route('/admin/export_excel')
@admin_required
def export_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        tasks = read_tasks(); slips = read_slips()
        wb = Workbook(); ws = wb.active; ws.title = 'Orders'
        headers = ['SN','ชื่องาน','ชื่อลูกค้า','เบอร์โทร','สถานะ','ความเร่งด่วน','กำหนดส่ง','วันที่สร้าง']
        for col, h in enumerate(headers, 1): ws.cell(row=1, column=col, value=h).font = Font(bold=True)
        for i, task in enumerate(tasks, 2):
            ws.cell(row=i, column=1, value=task.get('sn',''))
            ws.cell(row=i, column=2, value=task.get('title',''))
            ws.cell(row=i, column=3, value=task['customer']['name'])
            ws.cell(row=i, column=4, value=task['customer']['phone'])
            ws.cell(row=i, column=5, value=task.get('status',''))
            ws.cell(row=i, column=6, value=task.get('priority',''))
            ws.cell(row=i, column=7, value=task.get('deadline',''))
            ws.cell(row=i, column=8, value=task.get('createdAt','')[:10])
        buf = _io.BytesIO(); wb.save(buf)
        fname = f'orders_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        resp = make_response(buf.getvalue())
        resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        resp.headers['Content-Disposition'] = f'attachment; filename="{fname}"'
        return resp
    except Exception as e: return f'เกิดข้อผิดพลาด: {e}', 500

# ── Spec Sheet PDF ────────────────────────────────────────────────────────────
@app.route('/admin/spec_sheet/<task_id>')
@admin_required
def admin_spec_sheet(task_id):
    if not PDF_ENABLED or not generate_spec_sheet:
        return 'PDF ไม่พร้อมใช้งาน', 500
    tasks = read_tasks()
    task  = next((t for t in tasks if t['id'] == task_id), None)
    if not task: return 'ไม่พบ', 404
    pdf_bytes = generate_spec_sheet(task, COMPANY_NAME)
    resp = make_response(pdf_bytes)
    resp.headers['Content-Type'] = 'application/pdf'
    resp.headers['Content-Disposition'] = f'inline; filename="spec_{task_id[-6:]}.pdf"'
    return resp

# ── 3D File viewer (list files for a task) ────────────────────────────────────
@app.route('/admin/task_files/<task_id>')
@admin_required
def admin_task_files(task_id):
    tasks = read_tasks()
    task  = next((t for t in tasks if t['id'] == task_id), None)
    if not task: return jsonify({'error': 'not found'}), 404
    files = task.get('specs_3d', {}).get('files', []) or task.get('specs_custom', {}).get('reference_files', [])
    folder = MODEL_3D_FOLDER if task.get('specs_3d') else CUSTOM_ORDER_FOLDER
    url_part = '3d_models' if task.get('specs_3d') else 'custom_orders'
    # Build URLs
    file_list = []
    for sf in files:
        fname = sf.get('filename', '')
        path  = os.path.join(folder, fname)
        file_list.append({
            **sf,
            'url':    f'/uploads/{url_part}/{fname}',
            'exists': os.path.exists(path),
            'size':   os.path.getsize(path) if os.path.exists(path) else 0,
        })
    return jsonify({'task_id': task_id, 'title': task.get('title',''), 'files': file_list})

# ── Customer Accounts ──────────────────────────────────────────────────────────
def read_customers():  return _r('customers.json', {})
def write_customers(d): _w('customers.json', d)

_init('customers.json', {})

def _hash_pw(pw): return hash_password(pw)

@app.route('/customer/register', methods=['GET','POST'])
def customer_register():
    error = None; success = False
    if request.method == 'POST':
        name  = request.form.get('name','').strip()
        phone = request.form.get('phone','').strip()
        email = request.form.get('email','').strip().lower()
        pw    = request.form.get('password','')
        pw2   = request.form.get('password2','')
        customers = read_customers()
        if not name or not phone or not pw:
            error = 'กรุณากรอกข้อมูลให้ครบถ้วน'
        elif phone in customers:
            error = 'เบอร์โทรนี้ลงทะเบียนแล้ว กรุณาเข้าสู่ระบบ'
        elif len(pw) < 6:
            error = 'รหัสผ่านต้องมีอย่างน้อย 6 ตัวอักษร'
        elif pw != pw2:
            error = 'รหัสผ่านไม่ตรงกัน'
        else:
            customers[phone] = {
                'name': name, 'phone': phone, 'email': email,
                'password': _hash_pw(pw),
                'created_at': datetime.now().isoformat(),
            }
            write_customers(customers)
            session['customer_phone'] = phone
            session['customer_name']  = name
            return redirect('/customer/dashboard')
    return render_template('customer_register.html', error=error, active_page='register')

@app.route('/customer/login', methods=['GET','POST'])
def customer_login():
    error = None
    if request.method == 'POST':
        phone = request.form.get('phone','').strip()
        pw    = request.form.get('password','')
        customers = read_customers()
        c = customers.get(phone)
        if not c or not verify_password(c.get('password',''), pw):
            error = 'เบอร์โทรหรือรหัสผ่านไม่ถูกต้อง'
        else:
            if password_needs_upgrade(c.get('password','')):
                c['password'] = hash_password(pw)
                customers[phone] = c
                write_customers(customers)
            session['customer_phone'] = phone
            session['customer_name']  = c['name']
            return redirect('/customer/dashboard')
    return render_template('customer_login.html', error=error, active_page='clogin')

@app.route('/customer/logout')
def customer_logout():
    session.pop('customer_phone', None)
    session.pop('customer_name',  None)
    return redirect('/')

@app.route('/customer/dashboard')
def customer_dashboard():
    phone = session.get('customer_phone')
    if not phone: return redirect('/customer/login')
    tasks = [t for t in read_tasks() if t['customer']['phone'] == phone]
    tasks.sort(key=lambda t: t.get('createdAt',''), reverse=True)
    customers = read_customers()
    customer  = customers.get(phone, {})
    slips     = read_slips()
    # Attach slip info
    tasks_with_info = []
    for t in tasks:
        tc = dict(t)
        tc['slips'] = slips.get(t['id'], [])
        tasks_with_info.append(tc)
    return render_template('customer_dashboard.html',
                           customer=customer, tasks=tasks_with_info,
                           status_labels=STATUS_LABELS, status_progress=STATUS_PROGRESS,
                           active_page='customer_dash')

@app.route('/customer/profile', methods=['GET','POST'])
def customer_profile():
    phone = session.get('customer_phone')
    if not phone: return redirect('/customer/login')
    customers = read_customers()
    customer  = customers.get(phone, {})
    message   = None; error = None
    if request.method == 'POST':
        name  = request.form.get('name','').strip()
        email = request.form.get('email','').strip()
        pw    = request.form.get('new_password','')
        pw2   = request.form.get('new_password2','')
        if name:
            customer['name']  = name
            session['customer_name'] = name
        if email:
            customer['email'] = email
        if pw:
            if len(pw) < 6:
                error = 'รหัสผ่านต้องมีอย่างน้อย 6 ตัวอักษร'
            elif pw != pw2:
                error = 'รหัสผ่านใหม่ไม่ตรงกัน'
            else:
                customer['password'] = _hash_pw(pw)
                message = 'อัปเดตข้อมูลเรียบร้อย'
        if not error:
            customers[phone] = customer
            write_customers(customers)
            message = message or 'อัปเดตข้อมูลเรียบร้อย'
    return render_template('customer_profile.html', customer=customer,
                           message=message, error=error, active_page='customer_profile')

# ── Startup ────────────────────────────────────────────────────────────────────
with app.app_context():
    backfill_sn()

if __name__ == '__main__':
    if os.path.exists('.env'):
        with open('.env', encoding='utf-8') as f:
            for line in f:
                if '=' in line and not line.startswith('#'):
                    k, v = line.strip().split('=', 1)
                    os.environ.setdefault(k, v)
    app.run(debug=True, port=5000)

