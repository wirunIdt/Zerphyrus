"""
app.py — ระบบจัดการงานลูกค้า v4
Features: Orders, Tickets, Stamp Cards, LINE Bot, PromptPay QR Upload,
          Payment Slip Verification, Queue Management, Work Calendar & Analytics
"""

import json, os, uuid
from datetime import datetime, date, timedelta
from functools import wraps
from collections import defaultdict

from flask import (Flask, render_template, request, redirect, url_for,
                   session, jsonify, make_response, send_from_directory)

from promptpay import generate_promptpay_payload
from pdf_generator import generate_order_pdf
from queue_manager import (
    read_queue, read_calendar, write_calendar,
    sync_queue, reorder_queue, set_task_estimate,
    get_queue_with_tasks, yearly_analytics,
    add_custom_date, remove_custom_date, update_calendar_settings,
    working_days_count, MONTH_TH
)

try:
    from line_handler import handle_events, verify_signature
    LINE_ENABLED = True
except Exception:
    LINE_ENABLED = False

# ── Fix .env encoding for Windows (prevent UnicodeDecodeError on startup) ──────
import sys
def _fix_env_encoding():
    """Re-save .env as UTF-8 if it was saved with wrong encoding (e.g. cp1252 on Windows)."""
    env_path = os.path.join(os.path.dirname(__file__), '.env')
    if not os.path.exists(env_path):
        return
    for enc in ('utf-8', 'utf-8-sig', 'cp1252', 'tis-620', 'latin-1'):
        try:
            with open(env_path, 'r', encoding=enc) as f:
                text = f.read()
            # Re-write as clean UTF-8 (no BOM)
            with open(env_path, 'w', encoding='utf-8') as f:
                f.write(text)
            break
        except (UnicodeDecodeError, LookupError):
            continue
_fix_env_encoding()


# ── App setup ──────────────────────────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'change-this-secret-in-production')

# Allow HTTPS behind proxy (ngrok, nginx, etc.)
from werkzeug.middleware.proxy_fix import ProxyFix
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)

PREFERRED_SCHEME = os.environ.get('PREFERRED_SCHEME', '')   # set to 'https' if behind proxy

UPLOAD_FOLDER = 'uploads'
QR_FOLDER     = os.path.join(UPLOAD_FOLDER, 'qr')
SLIP_FOLDER   = os.path.join(UPLOAD_FOLDER, 'slips')
ALLOWED_IMG   = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

for d in [UPLOAD_FOLDER, QR_FOLDER, SLIP_FOLDER]:
    os.makedirs(d, exist_ok=True)

PROMPTPAY_PHONE  = os.environ.get('PROMPTPAY_PHONE', '0812345678')
COMPANY_NAME     = os.environ.get('COMPANY_NAME', 'ระบบจัดการงานลูกค้า')

@app.context_processor
def inject_globals():
    """Inject cart_count and company_name into every template automatically."""
    try:
        cc = cart_count()
    except Exception:
        cc = 0
    return dict(cart_count=cc, company_name=COMPANY_NAME)

STAMPS_TO_REWARD = 10

# ── Data helpers ───────────────────────────────────────────────────────────────
def _init(path, default):
    if not os.path.exists(path):
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(default, f, ensure_ascii=False)

for p, d in [
    ('tasks.json', []),
    ('users.json', {'admin': 'admin123'}),
    ('stamps.json', {}),
    ('tickets.json', {}),
    ('slips.json', {}),          # {task_id: [{slip_file, uploaded_at, status, note}]}
    ('products.json', []),        # product catalog
    ('orders_cart.json', {}),     # {session_id: [{product_id, qty, ...}]}
    ('sn_counter.json', {'last_sn': 0}),   # auto-increment SN
]:
    _init(p, d)

def _r(p, default):
    try:
        with open(p, 'r', encoding='utf-8') as f: return json.load(f)
    except: return default

def _w(p, data):
    with open(p, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

read_tasks    = lambda: _r('tasks.json', [])
write_tasks   = lambda d: _w('tasks.json', d)
read_users    = lambda: _r('users.json', {'admin': 'admin123'})
write_users   = lambda d: _w('users.json', d)
read_stamps   = lambda: _r('stamps.json', {})
write_stamps  = lambda d: _w('stamps.json', d)
read_tickets  = lambda: _r('tickets.json', {})
write_tickets = lambda d: _w('tickets.json', d)
read_slips    = lambda: _r('slips.json', {})
write_slips   = lambda d: _w('slips.json', d)
read_products = lambda: _r('products.json', [])
read_todos    = lambda: _r('todos.json', [])
write_todos   = lambda d: _w('todos.json', d)
write_products= lambda d: _w('products.json', d)
read_sn       = lambda: _r('sn_counter.json', {'last_sn': 0})
write_sn      = lambda d: _w('sn_counter.json', d)

def next_sn() -> str:
    """Generate next SN in format ORD-YYYYMM-XXXX (e.g. ORD-202602-0001)."""
    counter = read_sn()
    counter['last_sn'] = counter.get('last_sn', 0) + 1
    write_sn(counter)
    prefix = datetime.now().strftime('%Y%m')
    return f"ORD-{prefix}-{counter['last_sn']:04d}"

def backfill_sn():
    """Add SN to existing tasks that don't have one (migration helper)."""
    tasks = read_tasks()
    counter = read_sn()
    changed = False
    for t in reversed(tasks):   # oldest first
        if not t.get('sn'):
            counter['last_sn'] = counter.get('last_sn', 0) + 1
            t['sn'] = f"ORD-MIGR-{counter['last_sn']:04d}"
            changed = True
    if changed:
        write_sn(counter)
        write_tasks(tasks)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_IMG

def get_qr_image():
    """Return filename of current PromptPay QR image (if uploaded)."""
    for ext in ALLOWED_IMG:
        path = os.path.join(QR_FOLDER, f'promptpay.{ext}')
        if os.path.exists(path):
            return f'promptpay.{ext}'
    return None

def add_stamp(phone, name=''):
    s = read_stamps()
    if phone not in s:
        s[phone] = {'stamps': 0, 'total_earned': 0, 'rewards_redeemed': 0, 'name': name}
    s[phone]['stamps'] += 1
    s[phone]['total_earned'] += 1
    write_stamps(s)

def create_ticket(task):
    code = uuid.uuid4().hex[:8].upper()
    t = read_tickets()
    t[code] = {
        'task_id': task['id'],
        'customer_name': task['customer']['name'],
        'customer_phone': task['customer']['phone'],
        'task_title': task['title'],
        'status': 'active',
        'created_at': datetime.now().isoformat(),
        'checked_in_at': None,
        'checked_in_by': None,
    }
    write_tickets(t)
    return code

def build_analytics(tasks):
    sc = defaultdict(int); pc = defaultdict(int)
    for t in tasks:
        sc[t['status']] += 1
        pc[t.get('priority', 'medium')] += 1
    today = date.today()
    days = [(today - timedelta(days=i)).isoformat() for i in range(13, -1, -1)]
    dc = defaultdict(int)
    for t in tasks: dc[t['createdAt'][:10]] += 1
    n = len(tasks); c = sc.get('completed', 0)
    return {
        'status_labels': ['รอดำเนินการ', 'กำลังทำ', 'เสร็จสิ้น', 'ยกเลิก'],
        'status_values': [sc.get(k, 0) for k in ['pending', 'inprogress', 'completed', 'cancelled']],
        'priority_labels': ['เร่งด่วน', 'ปานกลาง', 'ไม่เร่งด่วน'],
        'priority_values': [pc.get(k, 0) for k in ['high', 'medium', 'low']],
        'day_labels': days,
        'day_values': [dc[d] for d in days],
        'completion_rate': round(c / n * 100, 1) if n else 0,
    }

def slip_status_for_task(task_id):
    """Get latest slip status for a task."""
    slips = read_slips()
    task_slips = slips.get(task_id, [])
    if not task_slips:
        return None
    return task_slips[-1]   # latest slip

def pending_slips_count():
    slips = read_slips()
    count = 0
    for task_slips in slips.values():
        for s in task_slips:
            if s.get('status') == 'pending':
                count += 1
    return count

def get_webhook_url():
    """Build webhook URL - HTTPS if behind proxy or env set."""
    base = request.url_root.rstrip('/')
    if PREFERRED_SCHEME:
        base = base.replace('http://', f'{PREFERRED_SCHEME}://')
    elif request.headers.get('X-Forwarded-Proto') == 'https':
        base = base.replace('http://', 'https://')
    return f"{base}/webhook"

# ── Context builder ────────────────────────────────────────────────────────────
def admin_context(tasks_override=None):
    all_tasks = read_tasks()
    tasks     = tasks_override if tasks_override is not None else all_tasks
    stamps    = read_stamps()
    tickets   = read_tickets()
    cal       = read_calendar()
    yr        = date.today().year
    slips     = read_slips()
    # Attach slip info to tasks
    tasks_with_slip = []
    for t in tasks:
        tc = dict(t)
        tc['slip'] = slip_status_for_task(t['id'])
        tasks_with_slip.append(tc)

    return dict(
        tasks            = tasks_with_slip,
        username         = session.get('username', ''),
        stats            = {
            'total':     len(all_tasks),
            'pending':   sum(1 for t in all_tasks if t['status'] == 'pending'),
            'inprogress':sum(1 for t in all_tasks if t['status'] == 'inprogress'),
            'completed': sum(1 for t in all_tasks if t['status'] == 'completed'),
        },
        analytics        = build_analytics(all_tasks),
        stamps           = stamps,
        stamp_stats      = {
            'total_customers': len(stamps),
            'total_stamps':    sum(v['stamps'] for v in stamps.values()),
            'total_redeemed':  sum(v['rewards_redeemed'] for v in stamps.values()),
        },
        tickets          = tickets,
        ticket_stats     = {
            'total':      len(tickets),
            'active':     sum(1 for t in tickets.values() if t['status'] == 'active'),
            'checked_in': sum(1 for t in tickets.values() if t['status'] == 'checked_in'),
        },
        stamps_to_reward = STAMPS_TO_REWARD,
        promptpay_phone  = PROMPTPAY_PHONE,
        calendar         = cal,
        ya               = yearly_analytics(all_tasks, yr, cal),
        queue_tasks      = get_queue_with_tasks(all_tasks, cal),
        qr_image         = get_qr_image(),
        all_slips        = slips,
        pending_slips    = pending_slips_count(),
        todos            = read_todos(),
    )

# ── Auth ───────────────────────────────────────────────────────────────────────
def admin_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if 'username' not in session:
            return redirect(url_for('login'))
        return f(*a, **kw)
    return dec

# ══════════════════════════════════════════════════════════════════════════════
#  PUBLIC
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/')
def index():
    return render_template('order_form.html', company_name=COMPANY_NAME, cart_count=cart_count(), active_page='order')

@app.route('/submit_order', methods=['POST'])
def submit_order():
    tasks = read_tasks()
    task  = {
        'id': str(int(datetime.now().timestamp() * 1000)),
        'sn': next_sn(),
        'customer': {
            'name':  request.form.get('customer_name', ''),
            'phone': request.form.get('customer_phone', ''),
            'email': request.form.get('customer_email', ''),
        },
        'title':       request.form.get('task_title', ''),
        'description': request.form.get('task_description', ''),
        'priority':    request.form.get('priority', 'medium'),
        'deadline':    request.form.get('deadline', ''),
        'status':      'pending',
        'createdBy':   'ลูกค้า',
        'createdAt':   datetime.now().isoformat(),
        'updatedAt':   datetime.now().isoformat(),
    }
    tasks.insert(0, task)
    write_tasks(tasks)
    code = create_ticket(task)
    return render_template('order_form.html', success=True,
                           ticket_code=code,
                           customer_name=task['customer']['name'],
                           task_id=task['id'],
                           order_sn=task.get('sn',''),
                           company_name=COMPANY_NAME, cart_count=cart_count(), active_page='order')
@app.route('/model')
def model():
    tasks = read_tasks()
    task  = {
        'id': str(int(datetime.now().timestamp() * 1000)),
        'sn': next_sn(),
        'customer': {
            'name':  request.form.get('customer_name', ''),
            'phone': request.form.get('customer_phone', ''),
            'email': request.form.get('customer_email', ''),
        },
        'title':       request.form.get('task_title', ''),
        'description': request.form.get('task_description', ''),
        'priority':    request.form.get('priority', 'medium'),
        'deadline':    request.form.get('deadline', ''),
        'status':      'pending',
        'createdBy':   'ลูกค้า',
        'createdAt':   datetime.now().isoformat(),
        'updatedAt':   datetime.now().isoformat(),
    }
    tasks.insert(0, task)
    write_tasks(tasks)
    code = create_ticket(task)
    return render_template('model.html', success=True,
                           ticket_code=code,
                           customer_name=task['customer']['name'],
                           task_id=task['id'],
                           order_sn=task.get('sn',''),
                           company_name=COMPANY_NAME, cart_count=cart_count(), active_page='order')

@app.route('/tracking')
def tracking():
    q  = request.args.get('q', '').strip().lower()
    sn = request.form.get('search_name', '').lower() or q
    sd = request.args.get('date', '') or request.form.get('search_date', '')
    if q or sn or sd:
        tasks = read_tasks()
        r = tasks
        if sn:
            r = [t for t in r if
                 sn in t['customer']['name'].lower() or
                 sn in t['customer']['phone'].lower() or
                 sn in (t.get('sn') or '').lower() or
                 sn in t['title'].lower() or
                 any(sn in (tk or '').lower() for tk in [t.get('ticket_code','')])]
        if sd:
            r = [t for t in r if t.get('deadline','').startswith(sd)]
        return render_template('tracking.html', results=r, searched=True,
                               company_name=COMPANY_NAME, cart_count=cart_count(),
                               active_page='tracking', search_q=q)
    return render_template('tracking.html', company_name=COMPANY_NAME,
                           cart_count=cart_count(), active_page='tracking', searched=False)

@app.route('/search', methods=['POST'])
def search():
    sn = request.form.get('search_name', '').lower()
    sd = request.form.get('search_date', '')
    tasks = read_tasks()
    r = tasks
    if sn:
        r = [t for t in r if
             sn in t['customer']['name'].lower() or
             sn in t['customer']['phone'].lower() or
             sn in (t.get('sn') or '').lower() or
             sn in t['title'].lower()]
    if sd:
        r = [t for t in r if t.get('deadline','').startswith(sd)]
    return render_template('tracking.html', results=r, searched=True,
                           company_name=COMPANY_NAME, cart_count=cart_count(),
                           active_page='tracking', search_q=sn)

# ── Serve uploaded files ───────────────────────────────────────────────────────
@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)

# ── Payment ────────────────────────────────────────────────────────────────────
@app.route('/payment/<task_id>')
def payment(task_id):
    tasks = read_tasks()
    task  = next((t for t in tasks if t['id'] == task_id), None)
    if not task: return 'ไม่พบออเดอร์', 404
    tickets     = read_tickets()
    code        = next((c for c, tk in tickets.items() if tk['task_id'] == task_id), '')
    amount      = request.args.get('amount', type=float)
    qr_image    = get_qr_image()
    # Only generate payload if no custom QR image
    payload     = generate_promptpay_payload(PROMPTPAY_PHONE, amount) if not qr_image else ''
    slip        = slip_status_for_task(task_id)
    return render_template('payment.html',
                           task=task, ticket_code=code,
                           promptpay_payload=payload,
                           promptpay_phone=PROMPTPAY_PHONE,
                           amount=amount,
                           qr_image=qr_image,
                           slip=slip)

# ── Slip upload (customer) ─────────────────────────────────────────────────────
@app.route('/upload_slip/<task_id>', methods=['POST'])
def upload_slip(task_id):
    tasks = read_tasks()
    task  = next((t for t in tasks if t['id'] == task_id), None)
    if not task: return 'ไม่พบออเดอร์', 404

    file = request.files.get('slip_file')
    if not file or file.filename == '':
        return redirect(url_for('payment', task_id=task_id) + '?error=no_file')
    if not allowed_file(file.filename):
        return redirect(url_for('payment', task_id=task_id) + '?error=bad_type')

    ext      = file.filename.rsplit('.', 1)[1].lower()
    fname    = f"{task_id}_{int(datetime.now().timestamp())}.{ext}"
    filepath = os.path.join(SLIP_FOLDER, fname)
    file.save(filepath)

    slips = read_slips()
    slips.setdefault(task_id, []).append({
        'file':        f'slips/{fname}',
        'uploaded_at': datetime.now().isoformat(),
        'status':      'pending',
        'note':        '',
        'amount':      request.form.get('amount', ''),
    })
    write_slips(slips)

    tickets = read_tickets()
    code    = next((c for c, tk in tickets.items() if tk['task_id'] == task_id), '')
    return render_template('payment.html',
                           task=task, ticket_code=code,
                           promptpay_payload='',
                           promptpay_phone=PROMPTPAY_PHONE,
                           amount=None,
                           qr_image=get_qr_image(),
                           slip=slips[task_id][-1],
                           slip_uploaded=True)

# ── PDF ────────────────────────────────────────────────────────────────────────
@app.route('/order_pdf/<task_id>')
def order_pdf(task_id):
    try:
        tasks  = read_tasks()
        task   = next((t for t in tasks if t['id'] == task_id), None)
        if not task: return 'ไม่พบออเดอร์', 404
        tickets = read_tickets()
        code    = next((c for c, tk in tickets.items() if tk['task_id'] == task_id), '')
        amount  = request.args.get('amount', type=float)
        qr_img  = get_qr_image()
        payload = generate_promptpay_payload(PROMPTPAY_PHONE, amount) if (amount and not qr_img) else ''
        pdf_bytes = generate_order_pdf(task, code, payload, COMPANY_NAME)
        resp = make_response(pdf_bytes)
        resp.headers['Content-Type']        = 'application/pdf'
        resp.headers['Content-Disposition'] = f'attachment; filename="order_{task_id[-6:]}.pdf"'
        return resp
    except Exception as e:
        app.logger.error(f'PDF error: {e}', exc_info=True)
        return f'เกิดข้อผิดพลาดในการสร้าง PDF: {e}', 500

@app.route('/admin/order_pdf/<task_id>')
@admin_required
def admin_order_pdf(task_id):
    tasks   = read_tasks()
    task    = next((t for t in tasks if t['id'] == task_id), None)
    if not task: return 'ไม่พบ', 404
    tickets = read_tickets()
    code    = next((c for c, tk in tickets.items() if tk['task_id'] == task_id), '')
    inc_qr  = request.args.get('qr', '0') == '1'
    amount  = request.args.get('amount', type=float)
    qr_img  = get_qr_image()
    payload = generate_promptpay_payload(PROMPTPAY_PHONE, amount) if (inc_qr and not qr_img) else ''
    pdf_bytes = generate_order_pdf(task, code, payload, COMPANY_NAME)
    resp = make_response(pdf_bytes)
    resp.headers['Content-Type']        = 'application/pdf'
    resp.headers['Content-Disposition'] = f'attachment; filename="order_{task_id[-6:]}.pdf"'
    return resp

# ── Ticket ─────────────────────────────────────────────────────────────────────
@app.route('/ticket/<code>')
def view_ticket(code):
    tickets = read_tickets()
    ticket  = tickets.get(code.upper())
    if not ticket: return render_template('ticket.html', error=True, code=code)
    tasks   = read_tasks()
    task    = next((t for t in tasks if t['id'] == ticket['task_id']), None)
    return render_template('ticket.html', ticket=ticket, code=code.upper(), task=task)

@app.route('/checkin', methods=['GET', 'POST'])
def public_checkin():
    msg = None; ticket = None; code = ''
    if request.method == 'POST':
        code    = request.form.get('code', '').strip().upper()
        tickets = read_tickets()
        if code not in tickets:
            msg = ('error', f'ไม่พบ Ticket รหัส {code}')
        elif tickets[code]['status'] == 'checked_in':
            msg    = ('warning', f'Check-in แล้วเมื่อ {tickets[code]["checked_in_at"][:16].replace("T"," ")}')
            ticket = tickets[code]
        else:
            tickets[code].update({
                'status': 'checked_in',
                'checked_in_at': datetime.now().isoformat(),
                'checked_in_by': 'self',
            })
            write_tickets(tickets)
            ticket = tickets[code]
            msg    = ('success', f'Check-in สำเร็จ! ยินดีต้อนรับ {ticket["customer_name"]} 🎉')
    return render_template('checkin.html', message=msg, ticket=ticket, code=code, company_name=COMPANY_NAME, cart_count=cart_count(), active_page='checkin')

# ── LINE Webhook ───────────────────────────────────────────────────────────────
@app.route('/webhook', methods=['GET', 'POST'])
def line_webhook():
    if request.method == 'GET':
        return jsonify({'status': 'LINE Webhook active', 'url': get_webhook_url()})
    if not LINE_ENABLED:
        return jsonify({'error': 'LINE not configured'}), 500
    body = request.get_data()
    sig  = request.headers.get('X-Line-Signature', '')
    if not verify_signature(body, sig):
        return jsonify({'error': 'Bad signature'}), 403
    try:
        handle_events(json.loads(body).get('events', []), read_tasks, read_tickets)
    except Exception as e:
        app.logger.error(f'LINE: {e}')
    return jsonify({'status': 'ok'})

# ── Auth ───────────────────────────────────────────────────────────────────────
@app.route('/login', methods=['GET', 'POST'])
def login():
    users    = read_users()
    ft       = len(users) == 0
    if request.method == 'POST':
        u = request.form.get('username', '')
        p = request.form.get('password', '')
        if ft:
            users[u] = p; write_users(users)
            session['username'] = u
            return redirect(url_for('admin_dashboard'))
        elif u in users and users[u] == p:
            session['username'] = u
            return redirect(url_for('admin_dashboard'))
        else:
            return render_template('login.html', company_name=COMPANY_NAME, cart_count=0, active_page='login',
                                   error='ชื่อผู้ใช้หรือรหัสผ่านไม่ถูกต้อง',
                                   first_time=ft)
    return render_template('login.html', first_time=ft, company_name=COMPANY_NAME, cart_count=0, active_page='login')

@app.route('/logout')
def logout():
    session.pop('username', None)
    return redirect(url_for('index'))

# ══════════════════════════════════════════════════════════════════════════════
#  ADMIN
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/admin')
@admin_required
def admin_dashboard():
    return render_template('admin_dashboard.html', products_count=len(read_products()), **admin_context())

@app.route('/admin/filter/<status>')
@admin_required
def filter_tasks(status):
    all_tasks = read_tasks()
    tasks = all_tasks if status == 'all' else [t for t in all_tasks if t['status'] == status]
    ctx = admin_context(tasks); ctx['current_filter'] = status
    return render_template('admin_dashboard.html', **ctx)

@app.route('/admin/update_status', methods=['POST'])
@admin_required
def update_status():
    task_id    = request.form.get('task_id', '').strip()
    new_status = request.form.get('new_status', '').strip()
    valid      = {'pending', 'inprogress', 'completed', 'cancelled'}
    if new_status not in valid:
        return jsonify({'status': 'error', 'error': 'invalid status'}), 400
    tasks = read_tasks()
    for t in tasks:
        if t['id'] == task_id:
            old_status = t['status']
            if new_status == 'completed' and old_status != 'completed':
                add_stamp(t['customer']['phone'], t['customer']['name'])
            t['status']    = new_status
            t['updatedAt'] = datetime.now().isoformat()
            t['updatedBy'] = session.get('username', '')
            break
    write_tasks(tasks)
    updated = next((t for t in tasks if t['id'] == task_id), None)
    return jsonify({'status': 'ok', 'task': updated})

@app.route('/admin/delete', methods=['POST'])
@admin_required
def delete_task():
    tid = request.form.get('task_id')
    write_tasks([t for t in read_tasks() if t['id'] != tid])
    return jsonify({'status': 'ok'})

# ── QR Image upload ────────────────────────────────────────────────────────────
@app.route('/admin/upload_qr', methods=['POST'])
@admin_required
def upload_qr():
    file = request.files.get('qr')
    if not file or not allowed_file(file.filename):
        return redirect(url_for('line_config') + '?error=bad_file')
    # Remove old QR images
    for ext in ALLOWED_IMG:
        old = os.path.join(QR_FOLDER, f'promptpay.{ext}')
        if os.path.exists(old): os.remove(old)
    ext  = file.filename.rsplit('.', 1)[1].lower()
    path = os.path.join(QR_FOLDER, f'promptpay.{ext}')
    file.save(path)
    return redirect(url_for('line_config') + '?qr_saved=1')

@app.route('/admin/delete_qr', methods=['POST'])
@admin_required
def delete_qr():
    for ext in ALLOWED_IMG:
        p = os.path.join(QR_FOLDER, f'promptpay.{ext}')
        if os.path.exists(p): os.remove(p)
    return redirect(url_for('line_config'))

# ── Slip verification ──────────────────────────────────────────────────────────
@app.route('/admin/verify_slip', methods=['POST'])
@admin_required
def verify_slip():
    task_id  = request.form.get('task_id', '')
    slip_idx = int(request.form.get('slip_idx', 0))
    action   = request.form.get('action', 'approve')   # approve | reject
    note     = request.form.get('note', '')
    slips    = read_slips()
    task_slips = slips.get(task_id, [])
    if 0 <= slip_idx < len(task_slips):
        task_slips[slip_idx]['status']      = 'approved' if action == 'approve' else 'rejected'
        task_slips[slip_idx]['note']        = note
        task_slips[slip_idx]['verified_at'] = datetime.now().isoformat()
        task_slips[slip_idx]['verified_by'] = session.get('username', '')
        # If approved, auto-advance task to inprogress
        if action == 'approve':
            tasks = read_tasks()
            for t in tasks:
                if t['id'] == task_id and t['status'] == 'pending':
                    t['status']    = 'inprogress'
                    t['updatedAt'] = datetime.now().isoformat()
            write_tasks(tasks)
    write_slips(slips)
    return jsonify({'status': 'ok', 'slip_status': task_slips[slip_idx]['status'] if 0 <= slip_idx < len(task_slips) else ''})

# ── Stamp ──────────────────────────────────────────────────────────────────────
@app.route('/admin/redeem_stamp', methods=['POST'])
@admin_required
def redeem_stamp():
    phone = request.form.get('phone', '')
    s     = read_stamps()
    if phone in s and s[phone]['stamps'] >= STAMPS_TO_REWARD:
        s[phone]['stamps']            -= STAMPS_TO_REWARD
        s[phone]['rewards_redeemed']  += 1
        write_stamps(s)
    s = read_stamps()
    return jsonify({'status': 'ok', 'stamps': s.get(phone, {}).get('stamps', 0), 'redeemed': s.get(phone, {}).get('rewards_redeemed', 0)})

@app.route('/admin/add_stamp_manual', methods=['POST'])
@admin_required
def add_stamp_manual():
    phone = request.form.get('phone', '').strip()
    name  = request.form.get('name', '').strip()
    if phone: add_stamp(phone, name)
    s = read_stamps()
    return jsonify({'status': 'ok', 'phone': phone, 'stamps': s.get(phone, {}).get('stamps', 0)})

# ── Tickets ────────────────────────────────────────────────────────────────────
@app.route('/admin/checkin_ticket', methods=['POST'])
@admin_required
def admin_checkin_ticket():
    code    = request.form.get('code', '').strip().upper()
    tickets = read_tickets()
    if code in tickets and tickets[code]['status'] == 'active':
        tickets[code].update({
            'status':        'checked_in',
            'checked_in_at': datetime.now().isoformat(),
            'checked_in_by': session.get('username', ''),
        })
        write_tickets(tickets)
    return jsonify({'status': 'ok', 'code': code})

# ── Queue ──────────────────────────────────────────────────────────────────────
@app.route('/admin/queue/reorder', methods=['POST'])
@admin_required
def api_reorder_queue():
    data = request.get_json(silent=True) or {}
    reorder_queue(data.get('order', []))
    return jsonify({'status': 'ok'})

@app.route('/admin/queue/estimate', methods=['POST'])
@admin_required
def api_set_estimate():
    set_task_estimate(
        request.form.get('task_id', ''),
        float(request.form.get('hours', 0) or 0),
        request.form.get('note', ''),
    )
    return jsonify({'status': 'ok'})

# ── Calendar ───────────────────────────────────────────────────────────────────
@app.route('/admin/calendar/settings', methods=['POST'])
@admin_required
def api_calendar_settings():
    work_days = [int(d) for d in request.form.getlist('work_days')]
    capacity  = int(request.form.get('capacity', 3))
    update_calendar_settings(work_days, capacity)
    return jsonify({'status': 'ok'})

@app.route('/admin/calendar/add_date', methods=['POST'])
@admin_required
def api_add_date():
    ds    = request.form.get('date', '')
    dtype = request.form.get('type', 'holiday')
    note  = request.form.get('note', '')
    if ds: add_custom_date(ds, dtype, note)
    return jsonify({'status': 'ok'})

@app.route('/admin/calendar/remove_date', methods=['POST'])
@admin_required
def api_remove_date():
    ds = request.form.get('date', '')
    if ds: remove_custom_date(ds)
    return jsonify({'status': 'ok'})

@app.route('/admin/api/yearly/<int:year>')
@admin_required
def api_yearly(year):
    return jsonify(yearly_analytics(read_tasks(), year, read_calendar()))

# ── LINE Config ────────────────────────────────────────────────────────────────
@app.route('/admin/line_config', methods=['GET', 'POST'])
@admin_required
def line_config():
    msg = None; qr_saved = request.args.get('qr_saved') == '1'
    if request.method == 'POST':
        lines = []
        for k in ['LINE_CHANNEL_ACCESS_TOKEN', 'LINE_CHANNEL_SECRET',
                  'ADMIN_LINE_USER_ID', 'PROMPTPAY_PHONE', 'COMPANY_NAME',
                  'PREFERRED_SCHEME']:
            v = request.form.get(k, '').strip()
            if v: os.environ[k] = v; lines.append(f"{k}={v}")
        if lines:
            with open('.env', 'w', encoding='utf-8') as f: f.write('\n'.join(lines) + '\n')
        msg = 'บันทึกเรียบร้อย — รีสตาร์ทเซิร์ฟเวอร์เพื่อให้มีผล'
    return render_template('line_config.html',
                           msg=msg, qr_saved=qr_saved,
                           qr_image=get_qr_image(),
                           token=os.environ.get('LINE_CHANNEL_ACCESS_TOKEN', ''),
                           secret=os.environ.get('LINE_CHANNEL_SECRET', ''),
                           admin_id=os.environ.get('ADMIN_LINE_USER_ID', ''),
                           promptpay=PROMPTPAY_PHONE,
                           company=COMPANY_NAME,
                           scheme=PREFERRED_SCHEME,
                           webhook_url=get_webhook_url())

# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT  (v5+)
# ══════════════════════════════════════════════════════════════════════════════
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import io as _io

# ── Colour palette (openpyxl uses 6-char hex without #) ───────────────────────
_XL_HEADER_BG   = '4F46E5'
_XL_HEADER_FG   = 'FFFFFF'
_XL_ALT_ROW     = 'F5F3FF'
_XL_PENDING_BG  = 'FEF3C7'
_XL_INPROG_BG   = 'DBEAFE'
_XL_DONE_BG     = 'D1FAE5'
_XL_CANCEL_BG   = 'FEE2E2'
_XL_TOTAL_BG    = 'EDE9FE'
_XL_HIGH_FG     = 'DC2626'
_XL_MED_FG      = 'D97706'
_XL_LOW_FG      = '059669'

_XL_STATUS_BG   = {'pending': _XL_PENDING_BG, 'inprogress': _XL_INPROG_BG,
                    'completed': _XL_DONE_BG,  'cancelled':  _XL_CANCEL_BG}
_XL_STATUS_TH   = {'pending':'รอดำเนินการ', 'inprogress':'กำลังทำ',
                    'completed':'เสร็จสิ้น', 'cancelled':'ยกเลิก'}
_XL_PRIORITY_TH = {'high':'เร่งด่วน', 'medium':'ปานกลาง', 'low':'ทั่วไป'}
_XL_PRIORITY_FG = {'high': _XL_HIGH_FG, 'medium': _XL_MED_FG, 'low': _XL_LOW_FG}

def _xl_thin_border():
    s = Side(style='thin', color='D1D5DB')
    return Border(left=s, right=s, top=s, bottom=s)

def _xl_hdr(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=True, color=_XL_HEADER_FG, name='Arial', size=10)
    c.fill      = PatternFill('solid', fgColor=_XL_HEADER_BG)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border    = _xl_thin_border()
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def _xl_cell(ws, row, col, value, align='left', bold=False, bg=None, fg='1F2937'):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name='Arial', size=9, bold=bold, color=fg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    c.border    = _xl_thin_border()
    if bg:
        c.fill = PatternFill('solid', fgColor=bg)
    return c

def build_excel(tasks: list, slips: dict) -> bytes:
    wb = Workbook()

    # ════════════════════════════════════════════════════════
    #  Sheet 1 — Orders
    # ════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Orders'
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = 'A3'

    # Title row
    ws1.merge_cells('A1:P1')
    c = ws1['A1']
    c.value     = f'รายงานคำสั่งงานทั้งหมด  |  ส่งออกเมื่อ {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    c.font      = Font(bold=True, size=13, name='Arial', color='4F46E5')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill      = PatternFill('solid', fgColor='EEF2FF')
    ws1.row_dimensions[1].height = 32

    # Header row
    headers = [
        ('SN',14),('รหัสออเดอร์',20),('ชื่องาน',28),
        ('ชื่อลูกค้า',20),('เบอร์โทร',16),('อีเมล',24),
        ('สถานะ',16),('ความเร่งด่วน',16),('กำหนดส่ง',14),
        ('วันที่สร้าง',18),('อัปเดตล่าสุด',18),('สร้างโดย',14),
        ('Ticket Code',16),('สลิป',14),('จำนวนเงิน',12),('หมายเหตุสลิป',20),
    ]
    for col, (label, width) in enumerate(headers, 1):
        _xl_hdr(ws1, 2, col, label, width)
    ws1.row_dimensions[2].height = 28

    # Data rows
    tickets_all = read_tickets()
    for i, task in enumerate(tasks):
        row    = i + 3
        bg     = _XL_STATUS_BG.get(task.get('status', ''), _XL_ALT_ROW if i%2==0 else None)
        prio   = task.get('priority', 'medium')
        p_fg   = _XL_PRIORITY_FG.get(prio, '1F2937')

        slip_list  = slips.get(task['id'], [])
        slip       = slip_list[-1] if slip_list else None
        slip_stat  = {'pending':'รอยืนยัน','approved':'อนุมัติ','rejected':'ปฏิเสธ'
                      }.get(slip['status'],'') if slip else '-'
        slip_amt   = slip.get('amount','') if slip else ''
        slip_note  = slip.get('note','')   if slip else ''
        code       = next((c for c,tk in tickets_all.items() if tk['task_id']==task['id']),'')

        cols = [
            (task.get('sn','—'),                       'center', True,  bg,   '4F46E5'),
            (task.get('id','')[-10:],                  'center', False, bg,   '1F2937'),
            (task.get('title',''),                     'left',   True,  None, '1F2937'),
            (task['customer']['name'],                 'left',   False, bg,   '1F2937'),
            (task['customer']['phone'],                'center', False, bg,   '1F2937'),
            (task['customer'].get('email',''),         'left',   False, None, '1F2937'),
            (_XL_STATUS_TH.get(task.get('status',''),''), 'center', True, bg, '1F2937'),
            (_XL_PRIORITY_TH.get(prio,''),            'center', True,  bg,   p_fg),
            (task.get('deadline','') or '',            'center', False, None, '1F2937'),
            (task.get('createdAt','')[:10],            'center', False, bg,   '1F2937'),
            (task.get('updatedAt','')[:10],            'center', False, bg,   '1F2937'),
            (task.get('createdBy',''),                 'center', False, None, '1F2937'),
            (code,                                     'center', True,  None, '7C3AED'),
            (slip_stat,                                'center', False, None, '1F2937'),
            (str(slip_amt) if slip_amt else '',        'center', False, None, '1F2937'),
            (slip_note,                                'left',   False, None, '1F2937'),
        ]
        for col, (val, align, bold, cell_bg, fg) in enumerate(cols, 1):
            _xl_cell(ws1, row, col, val, align, bold, cell_bg, fg)
        ws1.row_dimensions[row].height = 20

    # Summary row
    last_r = len(tasks) + 3
    ws1.merge_cells(f'A{last_r}:F{last_r}')
    sc = ws1.cell(row=last_r, column=1, value=f'รวมทั้งหมด {len(tasks)} รายการ')
    sc.font      = Font(bold=True, size=10, name='Arial', color='4F46E5')
    sc.fill      = PatternFill('solid', fgColor=_XL_TOTAL_BG)
    sc.alignment = Alignment(horizontal='center', vertical='center')
    sc.border    = _xl_thin_border()
    for col in range(7, 17):
        ec = ws1.cell(row=last_r, column=col)
        ec.fill   = PatternFill('solid', fgColor=_XL_TOTAL_BG)
        ec.border = _xl_thin_border()

    ws1.auto_filter.ref = f'A2:{get_column_letter(len(headers))}{len(tasks)+2}'

    # ════════════════════════════════════════════════════════
    #  Sheet 2 — Summary
    # ════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Summary')
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells('A1:D1')
    h = ws2['A1']
    h.value     = 'สรุปสถิติออเดอร์'
    h.font      = Font(bold=True, size=13, name='Arial', color='4F46E5')
    h.alignment = Alignment(horizontal='center', vertical='center')
    h.fill      = PatternFill('solid', fgColor='EEF2FF')
    ws2.row_dimensions[1].height = 30

    total = len(tasks)

    # Status breakdown
    status_data = [
        ('รอดำเนินการ', sum(1 for t in tasks if t.get('status')=='pending'),    _XL_PENDING_BG),
        ('กำลังทำ',     sum(1 for t in tasks if t.get('status')=='inprogress'), _XL_INPROG_BG),
        ('เสร็จสิ้น',  sum(1 for t in tasks if t.get('status')=='completed'),  _XL_DONE_BG),
        ('ยกเลิก',      sum(1 for t in tasks if t.get('status')=='cancelled'),  _XL_CANCEL_BG),
    ]
    for col, lbl in enumerate(['สถานะ','จำนวน','%',''], 1):
        if lbl: _xl_hdr(ws2, 2, col, lbl, [22,12,14,0][col-1])
    ws2.row_dimensions[2].height = 24
    for i, (lbl, cnt, bg) in enumerate(status_data, 3):
        _xl_cell(ws2, i, 1, lbl,  'left',   True,  bg)
        _xl_cell(ws2, i, 2, cnt,  'center', True,  bg)
        _xl_cell(ws2, i, 3, f'{cnt/total*100:.1f}%' if total else '0%', 'center', False, bg)
        ws2.row_dimensions[i].height = 20
    _xl_cell(ws2, 7, 1, 'รวม',  'left',   True, _XL_TOTAL_BG)
    _xl_cell(ws2, 7, 2, total,  'center', True, _XL_TOTAL_BG)
    _xl_cell(ws2, 7, 3, '100%', 'center', True, _XL_TOTAL_BG)
    ws2.row_dimensions[7].height = 20

    # Priority breakdown
    prio_data = [
        ('เร่งด่วน', sum(1 for t in tasks if t.get('priority')=='high'),   _XL_CANCEL_BG),
        ('ปานกลาง', sum(1 for t in tasks if t.get('priority')=='medium'),  _XL_PENDING_BG),
        ('ทั่วไป',  sum(1 for t in tasks if t.get('priority','medium')=='low'), _XL_DONE_BG),
    ]
    for col, lbl in enumerate(['ความเร่งด่วน','จำนวน','%'], 1):
        _xl_hdr(ws2, 10, col, lbl)
    ws2.row_dimensions[10].height = 24
    for i, (lbl, cnt, bg) in enumerate(prio_data, 11):
        _xl_cell(ws2, i, 1, lbl,  'left',   True,  bg)
        _xl_cell(ws2, i, 2, cnt,  'center', True,  bg)
        _xl_cell(ws2, i, 3, f'{cnt/total*100:.1f}%' if total else '0%', 'center', False, bg)
        ws2.row_dimensions[i].height = 20

    # Monthly trend
    from collections import defaultdict as _dd
    monthly = _dd(int)
    for t in tasks:
        k = t.get('createdAt','')[:7]
        if k: monthly[k] += 1
    months = sorted(monthly)[-6:]

    for col, lbl in enumerate(['เดือน','จำนวนออเดอร์'], 1):
        _xl_hdr(ws2, 15, col, lbl, [16,18][col-1])
    ws2.row_dimensions[15].height = 24
    for i, m in enumerate(months, 16):
        _xl_cell(ws2, i, 1, m,           'center', True,  _XL_ALT_ROW)
        _xl_cell(ws2, i, 2, monthly[m],  'center', False, _XL_ALT_ROW)
        ws2.row_dimensions[i].height = 20

    # Bar chart
    if months:
        chart = BarChart()
        chart.type   = 'col'
        chart.title  = 'จำนวนออเดอร์รายเดือน'
        chart.style  = 10
        chart.width  = 16
        chart.height = 12
        data_ref = Reference(ws2, min_col=2, min_row=15, max_row=15+len(months))
        cats_ref = Reference(ws2, min_col=1, min_row=16, max_row=15+len(months))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws2.add_chart(chart, 'E2')

    # ════════════════════════════════════════════════════════
    #  Sheet 3 — Slips
    # ════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('Slips')
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells('A1:G1')
    sh = ws3['A1']
    sh.value     = 'บันทึกสลิปการชำระเงิน'
    sh.font      = Font(bold=True, size=13, name='Arial', color='059669')
    sh.alignment = Alignment(horizontal='center', vertical='center')
    sh.fill      = PatternFill('solid', fgColor='D1FAE5')
    ws3.row_dimensions[1].height = 30

    slip_headers = [('SN',14),('ชื่องาน',26),('ลูกค้า',20),
                    ('เบอร์โทร',16),('จำนวนเงิน',14),('สถานะ',16),('วันที่',20)]
    for col, (lbl, w) in enumerate(slip_headers, 1):
        _xl_hdr(ws3, 2, col, lbl, w)
    ws3.row_dimensions[2].height = 24

    slip_row = 3
    for task in tasks:
        for s in slips.get(task['id'], []):
            bg    = {'approved': _XL_DONE_BG, 'rejected': _XL_CANCEL_BG,
                     'pending':  _XL_PENDING_BG}.get(s.get('status',''), None)
            stat  = {'pending':'รอยืนยัน','approved':'อนุมัติ',
                     'rejected':'ปฏิเสธ'}.get(s.get('status',''),'-')
            vals  = [
                task.get('sn','—'), task.get('title',''),
                task['customer']['name'], task['customer']['phone'],
                str(s.get('amount','')) if s.get('amount') else '',
                stat,
                s.get('uploaded_at','')[:16].replace('T',' '),
            ]
            for col, val in enumerate(vals, 1):
                _xl_cell(ws3, slip_row, col, val,
                         'center' if col in [1,4,5,6,7] else 'left',
                         col == 6, bg)
            ws3.row_dimensions[slip_row].height = 20
            slip_row += 1

    if slip_row == 3:
        ws3.cell(row=3, column=1, value='ยังไม่มีสลิปที่บันทึก').font = Font(italic=True, color='9CA3AF')

    wb.active = ws1
    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

@app.route('/admin/export_excel')
@admin_required
def export_excel():
    try:
        tasks  = read_tasks()
        slips  = read_slips()
        xl     = build_excel(tasks, slips)
        fname  = f'orders_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        resp   = make_response(xl)
        resp.headers['Content-Type']        = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        resp.headers['Content-Disposition'] = f'attachment; filename="{fname}"'
        return resp
    except Exception as e:
        app.logger.error(f'Excel export error: {e}', exc_info=True)
        return f'เกิดข้อผิดพลาดในการสร้าง Excel: {e}', 500

# ══════════════════════════════════════════════════════════════════════════════
#  PRODUCT CATALOG  (v7)
# ══════════════════════════════════════════════════════════════════════════════
import secrets as _secrets

PRODUCT_IMG_FOLDER = os.path.join(UPLOAD_FOLDER, 'products')
os.makedirs(PRODUCT_IMG_FOLDER, exist_ok=True)

def _cart_key():
    """Get or create a cart session key."""
    if 'cart_id' not in session:
        session['cart_id'] = _secrets.token_hex(16)
    return session['cart_id']

def get_cart():
    cid = _cart_key()
    carts = _r('orders_cart.json', {})
    return carts.get(cid, [])

def save_cart(items):
    cid = _cart_key()
    carts = _r('orders_cart.json', {})
    carts[cid] = items
    _w('orders_cart.json', carts)

def cart_count():
    return sum(i['qty'] for i in get_cart())

def cart_total():
    products = read_products()
    pid_map  = {p['id']: p for p in products}
    return sum(pid_map.get(i['product_id'], {}).get('price', 0) * i['qty']
               for i in get_cart())

# ── Public catalog ─────────────────────────────────────────────────────────────
@app.route('/catalog')
def catalog():
    products = [p for p in read_products() if p.get('active', True)]
    category = request.args.get('cat', '')
    search   = request.args.get('q', '').lower()
    cats     = sorted({p.get('category','') for p in products if p.get('category')})
    if category:
        products = [p for p in products if p.get('category','') == category]
    if search:
        products = [p for p in products if search in p['name'].lower()
                    or search in p.get('description','').lower()]
    return render_template('catalog.html',
                           products=products, cart_count=cart_count(),
                           categories=cats, active_cat=category, search=search)

@app.route('/product/<pid>')
def product_detail(pid):
    products = read_products()
    product  = next((p for p in products if p['id'] == pid), None)
    if not product or not product.get('active', True):
        return redirect(url_for('catalog'))
    return render_template('product_detail.html',
                           product=product, cart_count=cart_count())

# ── Cart ───────────────────────────────────────────────────────────────────────
@app.route('/cart')
def view_cart():
    cart     = get_cart()
    products = read_products()
    pid_map  = {p['id']: p for p in products}
    items    = []
    total    = 0
    for item in cart:
        p = pid_map.get(item['product_id'])
        if p:
            subtotal = p['price'] * item['qty']
            total   += subtotal
            items.append({**item, 'product': p, 'subtotal': subtotal})
    return render_template('cart.html', items=items, total=total, cart_count=cart_count())

@app.route('/cart/add', methods=['POST'])
def cart_add():
    pid  = request.form.get('product_id', '')
    qty  = max(1, int(request.form.get('qty', 1)))
    back = request.form.get('back', url_for('catalog'))
    products = read_products()
    product  = next((p for p in products if p['id'] == pid and p.get('active',True)), None)
    if not product:
        return redirect(back)
    cart = get_cart()
    for item in cart:
        if item['product_id'] == pid:
            item['qty'] = min(item['qty'] + qty, product.get('stock', 9999))
            break
    else:
        cart.append({'product_id': pid, 'qty': qty})
    save_cart(cart)
    # If buy_now, go straight to checkout
    if request.form.get('buy_now'):
        return redirect(url_for('cart_checkout'))
    return redirect(back)

@app.route('/cart/update', methods=['POST'])
def cart_update():
    pid      = request.form.get('product_id','')
    qty      = int(request.form.get('qty', 0))
    products = read_products()
    product  = next((p for p in products if p['id'] == pid), None)
    cart     = get_cart()
    if qty <= 0:
        cart = [i for i in cart if i['product_id'] != pid]
    else:
        stock = product.get('stock') if product else None
        if stock is not None:
            qty = min(qty, stock)
        for item in cart:
            if item['product_id'] == pid:
                item['qty'] = qty; break
    save_cart(cart)
    return redirect(url_for('view_cart'))

@app.route('/cart/remove', methods=['POST'])
def cart_remove():
    pid  = request.form.get('product_id','')
    cart = [i for i in get_cart() if i['product_id'] != pid]
    save_cart(cart)
    return redirect(url_for('view_cart'))

@app.route('/cart/checkout', methods=['GET','POST'])
def cart_checkout():
    cart = get_cart()
    if not cart:
        return redirect(url_for('catalog'))

    products = read_products()
    pid_map  = {p['id']: p for p in products}
    items    = []
    total    = 0
    for item in cart:
        p = pid_map.get(item['product_id'])
        if p:
            subtotal = p['price'] * item['qty']
            total   += subtotal
            items.append({**item, 'product': p, 'subtotal': subtotal})

    if request.method == 'POST':
        name  = request.form.get('name','')
        phone = request.form.get('phone','')
        email = request.form.get('email','')
        addr  = request.form.get('address','')

        # Create one order task with all items
        tasks = read_tasks()
        desc  = 'สินค้า:\n' + '\n'.join(
            f'- {i["product"]["name"]} x{i["qty"]}  ฿{i["subtotal"]:.0f}'
            for i in items
        ) + f'\n\nที่อยู่จัดส่ง: {addr}'

        task = {
            'id':          str(int(datetime.now().timestamp() * 1000)),
            'sn':          next_sn(),
            'customer':    {'name': name, 'phone': phone, 'email': email},
            'title':       f'คำสั่งซื้อออนไลน์ ({len(items)} รายการ)',
            'description': desc,
            'priority':    'medium',
            'deadline':    '',
            'status':      'pending',
            'createdBy':   'ลูกค้า (cart)',
            'createdAt':   datetime.now().isoformat(),
            'updatedAt':   datetime.now().isoformat(),
            'order_total': total,
            'order_items': [{'pid': i['product_id'], 'name': i['product']['name'],
                             'price': i['product']['price'], 'qty': i['qty']} for i in items],
        }
        tasks.insert(0, task)
        write_tasks(tasks)
        ticket_code = create_ticket(task)
        save_cart([])  # clear cart

        # Reduce stock
        for item in items:
            for p in products:
                if p['id'] == item['product_id'] and p.get('stock') is not None:
                    p['stock'] = max(0, p['stock'] - item['qty'])
        write_products(products)

        return redirect(url_for('payment', task_id=task['id']) +
                        f'?amount={total:.0f}')

    return render_template('checkout.html',
                           items=items, total=total, cart_count=cart_count())

# ── Admin products ─────────────────────────────────────────────────────────────
@app.route('/admin/products', methods=['GET'])
@admin_required
def admin_products():
    return render_template('admin_products.html',
                           products=read_products(),
                           username=session.get('username',''))

@app.route('/admin/products/add', methods=['POST'])
@admin_required
def admin_product_add():
    products = read_products()
    pid      = str(int(datetime.now().timestamp() * 1000))
    product  = {
        'id':          pid,
        'name':        request.form.get('name','').strip(),
        'description': request.form.get('description','').strip(),
        'price':       float(request.form.get('price', 0) or 0),
        'category':    request.form.get('category','').strip(),
        'stock':       int(request.form.get('stock', 0) or 0),
        'active':      request.form.get('active') == 'on',
        'image':       '',
        'createdAt':   datetime.now().isoformat(),
    }
    # Handle image upload
    file = request.files.get('image')
    if file and file.filename and allowed_file(file.filename):
        ext  = file.filename.rsplit('.', 1)[1].lower()
        fname = f'{pid}.{ext}'
        file.save(os.path.join(PRODUCT_IMG_FOLDER, fname))
        product['image'] = f'products/{fname}'
    products.insert(0, product)
    write_products(products)
    return jsonify({'status': 'ok', 'id': pid, 'name': product['name']})

@app.route('/admin/products/edit/<pid>', methods=['POST'])
@admin_required
def admin_product_edit(pid):
    products = read_products()
    updated_name = ''
    for p in products:
        if p['id'] == pid:
            p['name']        = request.form.get('name', p['name']).strip()
            p['description'] = request.form.get('description', p.get('description','')).strip()
            try:
                p['price']   = float(request.form.get('price', p['price']) or 0)
            except (ValueError, TypeError):
                p['price']   = p.get('price', 0)
            try:
                stock_val    = request.form.get('stock', '')
                p['stock']   = int(stock_val) if stock_val.strip() else None
            except (ValueError, TypeError):
                p['stock']   = p.get('stock')
            p['category']    = request.form.get('category', p.get('category','')).strip()
            p['active']      = request.form.get('active', 'off') == 'on'
            updated_name     = p['name']
            # Image update
            file = request.files.get('image')
            if file and file.filename and allowed_file(file.filename):
                ext   = file.filename.rsplit('.', 1)[1].lower()
                fname = f'{pid}.{ext}'
                file.save(os.path.join(PRODUCT_IMG_FOLDER, fname))
                p['image'] = f'products/{fname}'
            break
    write_products(products)
    return jsonify({'status': 'ok', 'name': updated_name})

@app.route('/admin/products/delete/<pid>', methods=['POST'])
@admin_required
def admin_product_delete(pid):
    products = [p for p in read_products() if p['id'] != pid]
    write_products(products)
    # Remove image file
    for ext in ALLOWED_IMG:
        path = os.path.join(PRODUCT_IMG_FOLDER, f'{pid}.{ext}')
        if os.path.exists(path): os.remove(path)
    return jsonify({'status': 'ok'})

@app.route('/admin/products/toggle/<pid>', methods=['POST'])
@admin_required
def admin_product_toggle(pid):
    products = read_products()
    for p in products:
        if p['id'] == pid:
            p['active'] = not p.get('active', True)
            break
    write_products(products)
    return jsonify({'status': 'ok'})

# ── Contact Us ────────────────────────────────────────────────────────────────
@app.route('/contact', methods=['GET', 'POST'])
def contact():
    sent = False
    if request.method == 'POST':
        name    = request.form.get('name', '').strip()
        email   = request.form.get('email', '').strip()
        phone   = request.form.get('phone', '').strip()
        message = request.form.get('message', '').strip()
        if name and (email or phone):
            sent = True
    return render_template('contact.html', sent=sent, promptpay_phone=PROMPTPAY_PHONE,
                           company_name=COMPANY_NAME, cart_count=cart_count(), active_page='contact')

# ── Admin Todo routes ─────────────────────────────────────────────────────────
@app.route('/admin/todos', methods=['GET'])
@admin_required
def admin_todos_get():
    return jsonify(read_todos())

@app.route('/admin/todos/add', methods=['POST'])
@admin_required
def admin_todos_add():
    todos = read_todos()
    todo = {
        'id':        str(int(datetime.now().timestamp() * 1000)),
        'text':      request.form.get('text', '').strip(),
        'done':      False,
        'priority':  request.form.get('priority', 'medium'),
        'due':       request.form.get('due', ''),
        'createdAt': datetime.now().isoformat(),
    }
    if not todo['text']:
        return jsonify({'status': 'error', 'error': 'empty'}), 400
    todos.insert(0, todo)
    write_todos(todos)
    return jsonify({'status': 'ok', 'todo': todo})

@app.route('/admin/todos/toggle/<tid>', methods=['POST'])
@admin_required
def admin_todos_toggle(tid):
    todos = read_todos()
    for t in todos:
        if t['id'] == tid:
            t['done'] = not t['done']
            break
    write_todos(todos)
    return jsonify({'status': 'ok'})

@app.route('/admin/todos/delete/<tid>', methods=['POST'])
@admin_required
def admin_todos_delete(tid):
    todos = [t for t in read_todos() if t['id'] != tid]
    write_todos(todos)
    return jsonify({'status': 'ok'})



# ── Run ────────────────────────────────────────────────────────────────────────
# Backfill SN for existing orders (safe to run multiple times)
with app.app_context():
    backfill_sn()

if __name__ == '__main__':
    if os.path.exists('.env'):
        with open('.env', encoding='utf-8') as f:
            for line in f:
                if '=' in line and not line.startswith('#'):
                    k, v = line.strip().split('=', 1)
                    os.environ.setdefault(k, v)
    app.run(debug=True, port=5000)
